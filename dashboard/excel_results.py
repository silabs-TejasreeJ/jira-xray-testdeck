"""Parse automation Excel reports (e.g. Report N.xlsx OverAllStatus) for TestDeck import."""

from __future__ import annotations

import io
import re
from dataclasses import asdict, dataclass, field
from datetime import datetime
from typing import Any

from openpyxl import load_workbook

from .html_results import STATUS_RANK, html_result_to_xray, normalize_map_id

PREFERRED_SHEETS = ("OverAllStatus", "OverallStatus", "Overall Status", "Results")

# Reject TestDeck export sheets so plan dumps are not mistaken for run reports.
REJECT_SHEETS = {
    "cases",
    "test executions",
    "defects",
    "case defects",
    "references",
}

# Header aliases → canonical (automation report only; do not map Status/Case ID
# from TestDeck Cases exports — those produce misleading PASS counts).
HEADER_ALIASES = {
    "mastercaseid": "master_case_id",
    "master_case_id": "master_case_id",
    "master_case": "master_case_id",
    "result": "result",
    "time": "time",
    "timestamp": "time",
    "testcase": "test_case",
    "test_case": "test_case",
    "technology": "technology",
    "testplan": "test_plan",
    "test_plan": "test_plan",
    "explanation": "explanation",
    "keyerror": "key_error",
    "key_error": "key_error",
}


@dataclass
class ExcelResultRow:
    source: str
    master_case_id: str
    test_src_map_id: str
    result_text: str
    status: str
    test_case: str = ""
    technology: str = ""
    time: str = ""
    explanation: str = ""
    key_error: str = ""
    sort_key: float = 0.0

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class ParsedExcelReport:
    source: str
    sheet: str = ""
    rows: list[ExcelResultRow] = field(default_factory=list)
    by_map_id: dict[str, str] = field(default_factory=dict)
    winning_sources: dict[str, str] = field(default_factory=dict)
    winning_rows: dict[str, ExcelResultRow] = field(default_factory=dict)
    errors: list[str] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return {
            "source": self.source,
            "sheet": self.sheet,
            "row_count": len(self.rows),
            "mapped_count": len(self.by_map_id),
            "status_counts": _count_statuses(self.by_map_id.values()),
            "errors": self.errors,
        }


def excel_result_to_xray(result_text: str) -> str:
    return html_result_to_xray(result_text)


def _norm_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[\s\-]+", "_", text)
    return HEADER_ALIASES.get(text, text)


def _parse_time_key(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    if isinstance(value, datetime):
        return value.timestamp()
    if isinstance(value, (int, float)):
        # Excel serial date roughly; prefer as-is if already unix-like
        num = float(value)
        if num > 1_000_000_000:
            return num
        # Excel epoch days → approximate unix (Excel 1899-12-30)
        try:
            return (num - 25569) * 86400.0
        except Exception:
            return num
    text = str(value).strip()
    if not text:
        return 0.0
    for fmt in (
        "%Y-%m-%d %H:%M:%S.%f%z",
        "%Y-%m-%d %H:%M:%S%z",
        "%Y-%m-%d %H:%M:%S.%f",
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%dT%H:%M:%S.%f%z",
        "%Y-%m-%dT%H:%M:%S%z",
        "%Y-%m-%dT%H:%M:%S.%f",
        "%Y-%m-%dT%H:%M:%S",
    ):
        try:
            cleaned = text.replace("+00:00", "+0000") if "%z" in fmt else text
            return datetime.strptime(cleaned, fmt).timestamp()
        except ValueError:
            continue
    # Strip timezone suffix like +00:00 for fromisoformat
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).timestamp()
    except Exception:
        return 0.0


def _pick_sheet(wb) -> str:
    names = list(wb.sheetnames)
    for preferred in PREFERRED_SHEETS:
        for name in names:
            if name.strip().lower() == preferred.lower():
                return name
    # Fallback: first non-export sheet with MasterCaseId + Result
    for name in names:
        if name.strip().lower() in REJECT_SHEETS:
            continue
        ws = wb[name]
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not header:
            continue
        cols = {_norm_header(h) for h in header if h is not None}
        if "master_case_id" in cols and "result" in cols:
            return name
    return ""


def _count_statuses(statuses) -> dict[str, int]:
    counts: dict[str, int] = {}
    for status in statuses:
        counts[status] = counts.get(status, 0) + 1
    return counts


def parse_excel_bytes(data: bytes, source: str = "upload.xlsx") -> ParsedExcelReport:
    report = ParsedExcelReport(source=source or "upload.xlsx")
    if not data:
        report.errors.append("Empty Excel file")
        return report

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        report.errors.append(f"Could not read Excel: {exc}")
        return report

    sheet_name = _pick_sheet(wb)
    if not sheet_name:
        report.errors.append(
            "Need an OverAllStatus sheet with MasterCaseId + Result columns "
            "(automation report). TestDeck Cases/export workbooks are not valid input."
        )
        return report
    report.sheet = sheet_name
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        report.errors.append(f"Sheet '{sheet_name}' is empty")
        return report

    col_index: dict[str, int] = {}
    for idx, cell in enumerate(header_row):
        key = _norm_header(cell)
        if key and key not in col_index:
            col_index[key] = idx

    if "master_case_id" not in col_index or "result" not in col_index:
        report.errors.append(
            "Expected columns MasterCaseId and Result "
            f"(sheet '{sheet_name}'). Found: "
            + ", ".join(str(h) for h in header_row if h is not None)
            + ". Upload the automation OverAllStatus report, not a TestDeck export."
        )
        return report

    id_i = col_index["master_case_id"]
    result_i = col_index["result"]
    time_i = col_index.get("time")
    case_i = col_index.get("test_case")
    tech_i = col_index.get("technology")
    expl_i = col_index.get("explanation")
    keyerr_i = col_index.get("key_error")

    def _cell(raw_row, idx: int | None) -> str:
        if idx is None or idx >= len(raw_row):
            return ""
        val = raw_row[idx]
        if val is None:
            return ""
        text = str(val).strip()
        # Cap huge log blobs for triage payloads / Excel cells.
        if len(text) > 4000:
            return text[:4000] + "…"
        return text

    best: dict[str, ExcelResultRow] = {}
    for raw in rows_iter:
        if not raw:
            continue
        master_raw = raw[id_i] if id_i < len(raw) else None
        result_raw = raw[result_i] if result_i < len(raw) else None
        if master_raw is None or master_raw == "":
            continue
        map_id = normalize_map_id(str(master_raw))
        if not map_id:
            continue
        result_text = str(result_raw or "").strip()
        status = excel_result_to_xray(result_text)
        time_val = raw[time_i] if time_i is not None and time_i < len(raw) else ""
        sort_key = _parse_time_key(time_val)
        row = ExcelResultRow(
            source=source,
            master_case_id=str(master_raw).strip(),
            test_src_map_id=map_id,
            result_text=result_text,
            status=status,
            test_case=_cell(raw, case_i),
            technology=_cell(raw, tech_i),
            time=str(time_val or "").strip(),
            explanation=_cell(raw, expl_i),
            key_error=_cell(raw, keyerr_i),
            sort_key=sort_key,
        )
        report.rows.append(row)
        prev = best.get(map_id)
        # Later timestamp wins; if equal/missing time, later row wins.
        if prev is None or row.sort_key >= prev.sort_key:
            best[map_id] = row

    report.by_map_id = {mid: row.status for mid, row in best.items()}
    report.winning_rows = best
    report.winning_sources = {
        mid: f"{source}:{row.time}" if row.time else source for mid, row in best.items()
    }
    if not report.by_map_id and not report.errors:
        report.errors.append("No MasterCaseId / Result rows found")
    return report


def excel_triage_statuses(
    report: ParsedExcelReport,
) -> tuple[dict[str, str], dict[str, dict[str, str]]]:
    """Worst status wins for failure triage (FAIL over a later PASS row)."""
    merged: dict[str, str] = {}
    meta: dict[str, dict[str, str]] = {}
    for row in report.rows or []:
        map_id = row.test_src_map_id or ""
        if not map_id:
            continue
        status = row.status or "TODO"
        prev = merged.get(map_id)
        row_meta = {
            "test_name": row.test_case or "",
            "technology": row.technology or "",
            "explanation": row.explanation or "",
            "key_error": row.key_error or "",
            "source": (
                f"{report.source}:{row.time}" if row.time else (report.source or "")
            ),
            "failed_api": "",
            "failed_command": "",
            "status_code": "",
            "failure_context": "",
        }
        if prev is None or STATUS_RANK.get(status, 0) > STATUS_RANK.get(prev, 0):
            merged[map_id] = status
            meta[map_id] = row_meta
        elif (
            prev == status
            and row.key_error
            and not (meta.get(map_id) or {}).get("key_error")
        ):
            meta[map_id] = row_meta
    return merged, meta
