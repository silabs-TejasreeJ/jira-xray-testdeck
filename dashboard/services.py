"""Dashboard domain services: normalize Xray/Jira data into TestRail-like views."""

from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from typing import Any

from django.conf import settings
from django.core.cache import cache

from .cache_utils import remember_key
from .excel_results import excel_triage_statuses, parse_excel_bytes
from .html_results import (
    ParsedHtmlReport,
    collect_html_paths,
    dedupe_named_blobs,
    extract_unix_timestamp,
    merge_reports,
    merge_reports_chronological,
    merge_reports_chronological_detailed,
    merge_reports_detailed,
    normalize_map_id,
    parse_html_bytes,
    parse_zip_files,
)
from .jira_client import JiraClient, JiraError
from .testrun_fields import (
    SEED_OPTIONS,
    empty_field_catalog,
    extract_named_ids,
    is_valid_trcf_id,
    match_field_key,
    normalize_trcf_value,
)
from .xray_client import XrayClient

EDITABLE_STATUSES = ["TODO", "EXECUTING", "PASS", "FAIL", "BLOCKED", "ABORTED", "NA"]

# Fallback Technology options (Silabs) when Jira createmeta is unavailable.
SEED_TECHNOLOGIES = [
    "WLAN + BLE",
    "WLAN",
    "BLE",
    "Bluetooth",
    "Zigbee",
    "Z-Wave",
    "Thread",
    "Matter",
    "Wi-SUN",
    "Proprietary",
    "Sidewalk",
    "Amazon Sidewalk",
    "WiFi",
    "15.4",
    "Multiprotocol",
    "Coex",
]

# Common Silabs stack_name values seen on Xray Test Plans (e.g. SW_SQA_TE-28392).
SEED_STACK_NAMES = [
    "917_IoT_FreeRTOS",
]

# Common release_name values on Xray Test Plans.
SEED_RELEASE_NAMES = [
    "wc-4.1.1",
]


STATUS_ALIASES = {
    "PASS": "PASS",
    "PASSED": "PASS",
    "FAIL": "FAIL",
    "FAILED": "FAIL",
    "ERROR": "FAIL",
    "EXECUTING": "EXECUTING",
    "TODO": "TODO",
    "UNTESTED": "TODO",
    "ABORTED": "ABORTED",
    "BLOCKED": "BLOCKED",
    "RETEST": "RETEST",
    "NA": "NA",
    "N/A": "NA",
    "NOT APPLICABLE": "NA",
}

STATUS_COLORS = {
    "PASS": "#70ad47",
    "FAIL": "#c00000",
    "TODO": "#b0b0b0",
    "BLOCKED": "#595959",
    "RETEST": "#ffc000",
    "EXECUTING": "#5b9bd5",
    "ABORTED": "#833c0c",
    "NA": "#7f7f7f",
    "OTHER": "#9e9e9e",
}


@dataclass
class StatusSummary:
    counts: dict[str, int]
    total: int
    passed: int
    failed: int
    todo: int
    pass_pct: float
    todo_pct: float
    chart: list[dict[str, Any]] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class SectionNode:
    id: str
    name: str
    path: str
    count: int = 0
    children: list["SectionNode"] = field(default_factory=list)
    status: StatusSummary | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "id": self.id,
            "name": self.name,
            "path": self.path,
            "count": self.count,
            "children": [c.to_dict() for c in self.children],
            "status": self.status.to_dict() if self.status else None,
        }


class DashboardService:
    def __init__(self) -> None:
        self.jira = JiraClient()
        self.xray = XrayClient(self.jira)
        self._field_cache: dict[str, str] | None = None

    def credentials_configured(self) -> bool:
        return self.jira.configured

    def connection_status(self) -> dict[str, Any]:
        if not self.jira.configured:
            return {
                "ok": False,
                "message": "Missing JIRA_USERNAME / JIRA_PASSWORD in .env",
            }
        try:
            me = self.jira.get_myself()
            return {
                "ok": True,
                "message": f"Connected as {me.get('displayName') or me.get('name')}",
                "user": me.get("displayName") or me.get("name"),
            }
        except JiraError as exc:
            return {
                "ok": False,
                "message": str(exc),
                "status_code": exc.status_code,
                "payload": exc.payload,
            }

    def list_technology_options(self, force_refresh: bool = False) -> list[str]:
        """Return Technology select-list values from Jira (cached)."""
        cache_key = "technology_options_v1"
        if not force_refresh:
            cached = cache.get(cache_key)
            if isinstance(cached, list) and cached:
                return cached

        options: list[str] = []
        seen: set[str] = set()

        def _add(value: Any) -> None:
            text = str(value or "").strip()
            if not text or text in seen:
                return
            seen.add(text)
            options.append(text)

        for seed in SEED_TECHNOLOGIES:
            _add(seed)

        field_id = (settings.XRAY_FIELD_MAP.get("technology") or "customfield_22640").strip()
        project = settings.JIRA_TEST_PROJECT_KEY or settings.JIRA_PROJECT_KEY
        issue_type = settings.JIRA_ISSUE_TYPE_TEST

        try:
            meta = self.jira.get(
                "/rest/api/2/issue/createmeta",
                params={
                    "projectKeys": project,
                    "issuetypeNames": issue_type,
                    "expand": "projects.issuetypes.fields",
                },
            )
            projects = meta.get("projects") if isinstance(meta, dict) else None
            for proj in projects or []:
                for itype in proj.get("issuetypes") or []:
                    fields = itype.get("fields") or {}
                    field = fields.get(field_id) or {}
                    for allowed in field.get("allowedValues") or []:
                        if isinstance(allowed, dict):
                            _add(allowed.get("value") or allowed.get("name"))
                        else:
                            _add(allowed)
        except JiraError:
            pass

        # Prefer default first in the list for UX.
        default = settings.DEFAULT_TECHNOLOGY
        if default and default in options:
            options = [default] + [o for o in options if o != default]

        cache.set(cache_key, options, max(settings.JIRA_CACHE_SECONDS, 1800))
        remember_key(cache_key)
        return options

    def list_stack_name_options(self, force_refresh: bool = False) -> list[str]:
        """Return stack_name values from Test Plan createmeta + recent plans."""
        cache_key = "stack_name_options_v1"
        if not force_refresh:
            cached = cache.get(cache_key)
            if isinstance(cached, list) and cached:
                return cached

        options: list[str] = []
        seen: set[str] = set()

        def _add(value: Any) -> None:
            text = str(value or "").strip()
            if not text or text in seen:
                return
            seen.add(text)
            options.append(text)

        for seed in SEED_STACK_NAMES:
            _add(seed)

        field_id = (
            settings.XRAY_FIELD_MAP.get("stack_name") or "customfield_32353"
        ).strip()
        project = settings.JIRA_PROJECT_KEY
        issue_type = settings.JIRA_ISSUE_TYPE_TEST_PLAN

        try:
            meta = self.jira.get(
                "/rest/api/2/issue/createmeta",
                params={
                    "projectKeys": project,
                    "issuetypeNames": issue_type,
                    "expand": "projects.issuetypes.fields",
                },
            )
            projects = meta.get("projects") if isinstance(meta, dict) else None
            for proj in projects or []:
                for itype in proj.get("issuetypes") or []:
                    fields = itype.get("fields") or {}
                    field = fields.get(field_id) or {}
                    for allowed in field.get("allowedValues") or []:
                        if isinstance(allowed, dict):
                            _add(allowed.get("value") or allowed.get("name"))
                        else:
                            _add(allowed)
        except JiraError:
            pass

        # Sample recent plans for free-text / historical stack values.
        try:
            jql = (
                f'project = "{project}" AND issuetype = "{issue_type}" '
                "ORDER BY updated DESC"
            )
            issues = self.jira.search(
                jql=jql,
                fields=["summary", field_id] if field_id else ["summary"],
                max_results=80,
            ).get("issues", [])
            for issue in issues:
                _add(self._extract_field(issue.get("fields") or {}, field_id))
        except JiraError:
            pass

        default = getattr(settings, "DEFAULT_STACK_NAME", "") or ""
        if default and default in options:
            options = [default] + [o for o in options if o != default]

        cache.set(cache_key, options, max(settings.JIRA_CACHE_SECONDS, 1800))
        remember_key(cache_key)
        return options

    def list_release_name_options(self, force_refresh: bool = False) -> list[str]:
        """Return release_name values from Test Plan createmeta + recent plans."""
        cache_key = "release_name_options_v1"
        if not force_refresh:
            cached = cache.get(cache_key)
            if isinstance(cached, list) and cached:
                return cached

        options: list[str] = []
        seen: set[str] = set()

        def _add(value: Any) -> None:
            text = str(value or "").strip()
            if not text or text in seen:
                return
            seen.add(text)
            options.append(text)

        for seed in SEED_RELEASE_NAMES:
            _add(seed)

        field_map = self.resolve_fields()
        field_id = (
            field_map.get("release_name")
            or settings.XRAY_FIELD_MAP.get("release_name")
            or ""
        ).strip()
        project = settings.JIRA_PROJECT_KEY
        issue_type = settings.JIRA_ISSUE_TYPE_TEST_PLAN

        if field_id:
            try:
                meta = self.jira.get(
                    "/rest/api/2/issue/createmeta",
                    params={
                        "projectKeys": project,
                        "issuetypeNames": issue_type,
                        "expand": "projects.issuetypes.fields",
                    },
                )
                projects = meta.get("projects") if isinstance(meta, dict) else None
                for proj in projects or []:
                    for itype in proj.get("issuetypes") or []:
                        fields = itype.get("fields") or {}
                        field = fields.get(field_id) or {}
                        for allowed in field.get("allowedValues") or []:
                            if isinstance(allowed, dict):
                                _add(allowed.get("value") or allowed.get("name"))
                            else:
                                _add(allowed)
            except JiraError:
                pass

            try:
                jql = (
                    f'project = "{project}" AND issuetype = "{issue_type}" '
                    "ORDER BY updated DESC"
                )
                issues = self.jira.search(
                    jql=jql,
                    fields=["summary", field_id],
                    max_results=80,
                ).get("issues", [])
                for issue in issues:
                    _add(self._extract_field(issue.get("fields") or {}, field_id))
            except JiraError:
                pass

        default = getattr(settings, "DEFAULT_RELEASE_NAME", "") or ""
        if default and default in options:
            options = [default] + [o for o in options if o != default]

        cache.set(cache_key, options, max(settings.JIRA_CACHE_SECONDS, 1800))
        remember_key(cache_key)
        return options

    @staticmethod
    def _jql_custom_equals(field_name: str, field_id: str | None, value: str) -> str:
        """Build a JQL equality clause for a custom field.

        Prefer cf[id] when known — bare names like stack_name often 400 on Jira
        Server unless quoted, and OR-ing both forms is a common failure mode.
        """
        safe = str(value).replace('"', '\\"').strip()
        if not safe:
            return ""
        fid = (field_id or "").strip()
        if fid.startswith("customfield_"):
            cf_num = fid.replace("customfield_", "")
            return f' AND cf[{cf_num}] = "{safe}"'
        return f' AND "{field_name}" = "{safe}"'

    @staticmethod
    def _jira_error_detail(exc: JiraError) -> str:
        payload = getattr(exc, "payload", None)
        if isinstance(payload, dict):
            msgs = payload.get("errorMessages") or []
            if isinstance(msgs, list) and msgs:
                return "; ".join(str(m) for m in msgs if m)
            errors = payload.get("errors") or {}
            if isinstance(errors, dict) and errors:
                return "; ".join(f"{k}: {v}" for k, v in errors.items())
        if payload:
            return str(payload)[:300]
        return str(exc)

    def resolve_fields(self) -> dict[str, str]:
        configured = {k: v for k, v in settings.XRAY_FIELD_MAP.items() if v}
        if self._field_cache is not None:
            # Keep auto-detected IDs, but always honor latest env/settings overrides.
            merged = dict(self._field_cache)
            merged.update(configured)
            self._field_cache = merged
            return self._field_cache

        # Cross-request cache: get_fields() is expensive and rarely changes.
        django_cache_key = "xray_field_map_resolved_v1"
        cached = cache.get(django_cache_key)
        if isinstance(cached, dict) and cached:
            merged = dict(cached)
            merged.update(configured)
            self._field_cache = merged
            return self._field_cache

        aliases = {
            "test_repo_path": [
                "test repository path",
                "testrepositorypath",
                "repository path",
                "xray test repository path",
            ],
            "test_environments": [
                "test environments",
                "test environment",
                "environments",
            ],
            "test_plan": ["test plan", "testplan"],
            "stack_name": ["stack_name", "stack name"],
            "release_name": ["release_name", "release name"],
            "feature_name": ["feature_name", "feature name"],
            "tech_area": ["tech_area", "tech area"],
            "testrail_section": [
                "testrail_section_id",
                "testrail section id",
                "testrail section",
            ],
            "technology": ["technology"],
            "test_src_map_id": ["test_src_map_id", "test src map id"],
            "case_id": ["case_id", "case id", "testrail case id", "testrail_case_id"],
        }

        try:
            fields = self.jira.get_fields()
        except JiraError:
            self._field_cache = configured
            return self._field_cache

        by_name: dict[str, str] = {}
        for item in fields:
            name = (item.get("name") or "").strip().lower()
            field_id = item.get("id")
            if name and field_id:
                by_name[name] = field_id

        resolved = dict(configured)
        for key, names in aliases.items():
            if resolved.get(key):
                continue
            for name in names:
                if name in by_name:
                    resolved[key] = by_name[name]
                    break

        self._field_cache = resolved
        cache.set(
            django_cache_key,
            resolved,
            max(int(getattr(settings, "JIRA_CACHE_SECONDS", 300) or 300), 6 * 3600),
        )
        remember_key(django_cache_key)
        return resolved

    def list_test_executions(self, query: str = "", limit: int = 40) -> list[dict[str, Any]]:
        project = settings.JIRA_PROJECT_KEY
        issue_type = settings.JIRA_ISSUE_TYPE_TEST_EXECUTION
        jql = (
            f'project = {project} AND issuetype = "{issue_type}" '
            f"ORDER BY updated DESC"
        )
        if query.strip():
            safe = query.replace('"', '\\"')
            if self._looks_like_issue_key(safe):
                clause = f'(key = "{safe}" OR summary ~ "{safe}")'
            else:
                clause = f'summary ~ "{safe}"'
            jql = (
                f'project = {project} AND issuetype = "{issue_type}" '
                f"AND {clause} ORDER BY updated DESC"
            )

        issues = self.jira.search(
            jql=jql,
            fields=["summary", "status", "updated", "created", "assignee", "reporter"],
            max_results=limit,
        ).get("issues", [])

        field_map = self.resolve_fields()
        results = []
        for issue in issues:
            fields = issue.get("fields") or {}
            results.append(
                {
                    "key": issue.get("key"),
                    "summary": fields.get("summary") or "",
                    "status": ((fields.get("status") or {}).get("name") or ""),
                    "updated": fields.get("updated"),
                    "created": fields.get("created"),
                    "assignee": self._user_name(fields.get("assignee")),
                    "reporter": self._user_name(fields.get("reporter")),
                    "url": self.jira.browse_url(issue["key"]),
                    "test_plan": self._extract_field(fields, field_map.get("test_plan")),
                    "environments": self._extract_field(
                        fields, field_map.get("test_environments")
                    ),
                }
            )
        return results

    def list_test_plans(
        self,
        query: str = "",
        limit: int = 40,
        technology: str | None = None,
        stack_name: str | None = None,
        release_name: str | None = None,
    ) -> list[dict[str, Any]]:
        """List Test Plans filtered by stack_name / release_name (not Technology)."""
        project = settings.JIRA_PROJECT_KEY
        issue_type = settings.JIRA_ISSUE_TYPE_TEST_PLAN
        # technology kept for call-site compat; plans use stack/release instead.
        _ = technology
        stack_name = (
            stack_name
            if stack_name is not None
            else getattr(settings, "DEFAULT_STACK_NAME", "") or ""
        )
        release_name = (
            release_name
            if release_name is not None
            else getattr(settings, "DEFAULT_RELEASE_NAME", "") or ""
        )
        field_map = self.resolve_fields()
        stack_fid = (
            field_map.get("stack_name")
            or settings.XRAY_FIELD_MAP.get("stack_name")
            or ""
        ).strip()
        release_fid = (
            field_map.get("release_name")
            or settings.XRAY_FIELD_MAP.get("release_name")
            or ""
        ).strip()

        def _build_jql(*, use_stack: bool, use_release: bool) -> str:
            jql = f'project = "{project}" AND issuetype = "{issue_type}"'
            if use_stack and stack_name and str(stack_name).strip():
                jql += self._jql_custom_equals(
                    "stack_name", stack_fid, str(stack_name)
                )
            if use_release and release_name and str(release_name).strip():
                # Skip release filter when field id is unknown — name-only often 400s.
                if release_fid.startswith("customfield_"):
                    jql += self._jql_custom_equals(
                        "release_name", release_fid, str(release_name)
                    )
                else:
                    jql += self._jql_custom_equals(
                        "release_name", None, str(release_name)
                    )
            if query.strip():
                safe = query.replace('"', '\\"')
                if self._looks_like_issue_key(safe):
                    jql += f' AND (key = "{safe}" OR summary ~ "{safe}")'
                else:
                    jql += f' AND summary ~ "{safe}"'
            return jql + " ORDER BY updated DESC"

        fields_core = ["summary", "status", "updated", "assignee"]
        fields_extra = [
            fid for fid in (stack_fid, release_fid) if fid.startswith("customfield_")
        ]
        fields_full = fields_core + [f for f in fields_extra if f not in fields_core]

        attempts: list[tuple[str, list[str]]] = [
            (
                _build_jql(use_stack=True, use_release=True),
                fields_full,
            ),
            (
                _build_jql(use_stack=True, use_release=True),
                fields_core,
            ),
            (
                _build_jql(use_stack=bool(stack_fid), use_release=False),
                fields_core,
            ),
            (
                _build_jql(use_stack=False, use_release=False),
                fields_core,
            ),
        ]

        issues: list[dict[str, Any]] = []
        last_exc: JiraError | None = None
        seen_jql: set[str] = set()
        for jql, fields in attempts:
            key = f"{jql}|{','.join(fields)}"
            if key in seen_jql:
                continue
            seen_jql.add(key)
            try:
                issues = self.jira.search(
                    jql=jql,
                    fields=fields,
                    max_results=limit,
                ).get("issues", [])
                last_exc = None
                break
            except JiraError as exc:
                last_exc = exc
                continue

        if last_exc is not None:
            detail = self._jira_error_detail(last_exc)
            raise JiraError(
                f"Unable to list Test Plans: {detail}",
                status_code=last_exc.status_code,
                payload=last_exc.payload,
            ) from last_exc

        return [
            {
                "key": issue.get("key"),
                "summary": (issue.get("fields") or {}).get("summary") or "",
                "status": (((issue.get("fields") or {}).get("status") or {}).get("name") or ""),
                "updated": (issue.get("fields") or {}).get("updated"),
                "assignee": self._user_name((issue.get("fields") or {}).get("assignee")),
                "stack_name": self._extract_field(
                    issue.get("fields") or {}, stack_fid or None
                ),
                "release_name": self._extract_field(
                    issue.get("fields") or {}, release_fid or None
                ),
                "url": self.jira.browse_url(issue["key"]),
            }
            for issue in issues
        ]

    def resolve_plan_ref(
        self,
        ref: str,
        technology: str | None = None,
        stack_name: str | None = None,
        release_name: str | None = None,
    ) -> dict[str, Any]:
        """Resolve a plan key or summary/name to a concrete Test Plan."""
        ref = (ref or "").strip()
        if not ref:
            raise JiraError("Test Plan key or name is required")
        stack_name = (
            stack_name
            if stack_name is not None
            else getattr(settings, "DEFAULT_STACK_NAME", "") or ""
        )
        release_name = (
            release_name
            if release_name is not None
            else getattr(settings, "DEFAULT_RELEASE_NAME", "") or ""
        )

        # Exact key — allow opening even if stack/release differ (explicit key wins).
        if self._looks_like_issue_key(ref):
            try:
                issue = self.jira.get_issue(ref, fields="summary,issuetype")
            except JiraError as exc:
                raise JiraError(f"Test Plan not found: {ref}") from exc
            fields = issue.get("fields") or {}
            return {
                "key": issue.get("key") or ref,
                "summary": fields.get("summary") or "",
                "url": self.jira.browse_url(issue.get("key") or ref),
            }

        # Search by summary / name, scoped by stack + release when set.
        matches: list[dict[str, Any]] = []
        seen_keys: set[str] = set()
        queries = [ref]
        token = ref.replace("_", " ").split()[0] if ref else ""
        if token and token != ref:
            queries.append(token)
        compact = "".join(ch if ch.isalnum() else " " for ch in ref).strip()
        if compact and compact not in queries:
            queries.append(compact.split()[0] if compact.split() else compact)

        for q in queries:
            try:
                for item in self.list_test_plans(
                    query=q,
                    limit=20,
                    stack_name=stack_name or "",
                    release_name=release_name or "",
                ):
                    key = item.get("key") or ""
                    if key and key not in seen_keys:
                        seen_keys.add(key)
                        matches.append(item)
            except JiraError:
                continue

        if not matches:
            try:
                recent = self.list_test_plans(
                    query="",
                    limit=80,
                    stack_name=stack_name or "",
                    release_name=release_name or "",
                )
            except JiraError:
                recent = []
            ref_l = ref.lower()
            matches = [
                m
                for m in recent
                if ref_l in (m.get("summary") or "").lower()
                or ref_l in (m.get("key") or "").lower()
            ]

        if not matches:
            notes = []
            if stack_name:
                notes.append(f'stack_name "{stack_name}"')
            if release_name:
                notes.append(f'release_name "{release_name}"')
            scope = f" for {', '.join(notes)}" if notes else ""
            raise JiraError(
                f'No Test Plan matched "{ref}"{scope}. '
                "Use a key like SW_SQA_TE-28392 or part of the plan summary."
            )
        exact = next(
            (m for m in matches if (m.get("summary") or "").strip().lower() == ref.lower()),
            None,
        )
        if not exact:
            exact = next(
                (
                    m
                    for m in matches
                    if ref.lower() in (m.get("summary") or "").lower()
                ),
                matches[0],
            )
        return exact

    def load_execution_cases(
        self,
        execution_key: str,
        technology: str | None = None,
        force_refresh: bool = False,
    ) -> list[dict[str, Any]]:
        """
        Fast path: parallel Xray pages (status + run id) + parallel Jira metadata.
        Cached to keep reloads snappy.
        """
        technology = technology if technology is not None else settings.DEFAULT_TECHNOLOGY
        tech_key = (technology or "ALL").replace(" ", "_").replace("+", "plus")
        cache_key = f"exec_cases_v7:{execution_key}:{tech_key}"
        if not force_refresh:
            cached = cache.get(cache_key)
            if cached is not None:
                return cached

        field_map = self.resolve_fields()

        def _fetch_runs() -> list[dict[str, Any]]:
            # Prefer small detailed pages (includes assignee). Fall back if Xray 500s.
            try:
                rows = self.xray.get_all_test_execution_tests(
                    execution_key,
                    detailed=True,
                    page_size=40,
                    hard_limit=settings.XRAY_HARD_LIMIT,
                )
                if rows and any(self._run_assignee(r) for r in rows[:30]):
                    return rows
            except JiraError:
                rows = []
            rows = self.xray.get_all_test_execution_tests(
                execution_key,
                detailed=False,
                hard_limit=settings.XRAY_HARD_LIMIT,
            )
            return self._enrich_run_assignees(rows)

        def _fetch_issues() -> list[dict[str, Any]]:
            if technology:
                safe_tech = technology.replace('"', '\\"')
                jql = (
                    f"issue in testExecutionTests({execution_key}) "
                    f'AND Technology = "{safe_tech}" ORDER BY key ASC'
                )
            else:
                jql = f"issue in testExecutionTests({execution_key}) ORDER BY key ASC"
            fields = [
                "summary",
                "description",
                "priority",
                "assignee",
                "labels",
                "components",
            ]
            for fid in field_map.values():
                if fid and fid not in fields:
                    fields.append(fid)
            return self.jira.search_all(
                jql=jql,
                fields=fields,
                page_size=200,
                hard_limit=settings.XRAY_HARD_LIMIT,
            )

        # Parallelize the two expensive fan-outs
        from concurrent.futures import ThreadPoolExecutor

        with ThreadPoolExecutor(max_workers=2) as pool:
            fut_runs = pool.submit(_fetch_runs)
            fut_issues = pool.submit(_fetch_issues)
            raw_tests = fut_runs.result()
            issues = fut_issues.result()

        status_by_key: dict[str, dict[str, Any]] = {}
        for item in raw_tests:
            key = (
                item.get("key")
                or item.get("testKey")
                or ((item.get("test") or {}).get("key") if isinstance(item.get("test"), dict) else None)
            )
            if not key:
                continue
            status_raw = item.get("status") or "TODO"
            if isinstance(status_raw, dict):
                status_raw = status_raw.get("name") or "TODO"
            display, user_key = self._run_assignee_info(item)
            status_by_key[key] = {
                "status": self._normalize_status(str(status_raw)),
                "run_id": item.get("id"),
                "assignee": display,
                "assignee_key": user_key,
                "defects": item.get("defects") or [],
                "rank": item.get("rank"),
            }

        cases: list[dict[str, Any]] = []
        for issue in issues:
            f = issue.get("fields") or {}
            key = issue.get("key")
            run = status_by_key.get(key, {})
            path = self._extract_field(f, field_map.get("test_repo_path")) or self._path_from_labels(
                f.get("labels") or []
            )
            # Prefer Test Run assignee (Xray "Assignee") over Test issue assignee.
            cases.append(
                {
                    "key": key,
                    "run_id": run.get("run_id"),
                    "summary": f.get("summary") or "",
                    "description": self._plain_description(f.get("description")),
                    "status": run.get("status", "TODO"),
                    "assignee": run.get("assignee")
                    or self._user_name(f.get("assignee"))
                    or "",
                    "assignee_key": run.get("assignee_key")
                    or self._user_key(f.get("assignee"))
                    or "",
                    "priority": ((f.get("priority") or {}).get("name") or ""),
                    "labels": f.get("labels") or [],
                    "components": [c.get("name") for c in (f.get("components") or [])],
                    "section_path": self._clean_path(path) or "Uncategorized",
                    "technology": self._extract_field(f, field_map.get("technology"))
                    or technology
                    or "",
                    "test_src_map_id": self._extract_field(
                        f, field_map.get("test_src_map_id")
                    ),
                    "case_id": self._extract_field(f, field_map.get("case_id")),
                    "testrail_section": self._extract_field(
                        f, field_map.get("testrail_section")
                    ),
                    "feature_name": self._extract_field(f, field_map.get("feature_name")),
                    "stack_name": self._extract_field(f, field_map.get("stack_name")),
                    "references": self._extract_field(f, field_map.get("references")),
                    "defects": run.get("defects") or [],
                    "rank": run.get("rank"),
                    "url": self.jira.browse_url(key),
                }
            )

        cache.set(cache_key, cases, settings.JIRA_CACHE_SECONDS)
        remember_key(cache_key)
        return cases

    @staticmethod
    def paginate_cases(
        cases: list[dict[str, Any]], page: int = 1, page_size: int | None = None
    ) -> dict[str, Any]:
        page_size = page_size or settings.UI_PAGE_SIZE
        page = max(1, int(page or 1))
        total = len(cases)
        total_pages = max(1, (total + page_size - 1) // page_size)
        if page > total_pages:
            page = total_pages
        start = (page - 1) * page_size
        end = start + page_size
        return {
            "page": page,
            "page_size": page_size,
            "total_pages": total_pages,
            "page_cases": cases[start:end],
            "showing_from": 0 if total == 0 else start + 1,
            "showing_to": min(end, total),
        }

    def _bust_execution_case_cache(self, execution_key: str) -> None:
        if not execution_key:
            return
        for tech in {settings.DEFAULT_TECHNOLOGY, "", "ALL", "WLAN + BLE"}:
            tech_key = (tech or "ALL").replace(" ", "_").replace("+", "plus")
            for prefix in (
                "exec_cases:",
                "exec_cases_v2:",
                "exec_cases_v3:",
                "exec_cases_v4:",
                "exec_cases_v5:",
                "exec_cases_v6:",
                "exec_cases_v7:",
            ):
                cache.delete(f"{prefix}{execution_key}:{tech_key}")

    def _normalize_defect_keys(self, defects: Any) -> list[str]:
        import re

        raw: list[str] = []
        if isinstance(defects, str):
            raw = re.split(r"[\s,;]+", defects)
        elif isinstance(defects, (list, tuple)):
            for item in defects:
                if isinstance(item, str):
                    raw.extend(re.split(r"[\s,;]+", item))
                elif item is not None:
                    raw.append(str(item))
        keys: list[str] = []
        for part in raw:
            text = (part or "").strip()
            if not text:
                continue
            # Allow paste of browse URL → extract issue key/ID.
            m = re.search(
                r"(?:/browse/|/issues/)?([A-Za-z][A-Za-z0-9_]+-\d+)\b",
                text,
            )
            key = (m.group(1) if m else text).strip().upper()
            if key and self._looks_like_issue_key(key) and key not in keys:
                keys.append(key)
        return keys

    def update_case_status(
        self,
        run_id: int | str,
        status: str,
        execution_key: str = "",
        defects: Any = None,
        custom_fields: Any = None,
    ) -> dict[str, Any]:
        status = self._normalize_status(status)
        if status not in EDITABLE_STATUSES and status not in STATUS_ALIASES.values():
            raise JiraError(f"Unsupported status: {status}")
        defect_keys = self._normalize_defect_keys(defects)
        cf_payload: list[dict[str, Any]] = []
        cf_skipped: list[str] = []
        if isinstance(custom_fields, list):
            cf_payload = XrayClient._normalize_custom_fields(custom_fields)
        elif isinstance(custom_fields, dict) and custom_fields:
            catalog = (
                self.get_test_run_custom_field_catalog(execution_key).get("fields")
                if execution_key
                else empty_field_catalog()
            )
            cf_payload, cf_skipped = self.build_custom_field_payload(
                custom_fields, catalog
            )
        self.xray.update_test_run(
            run_id, status=status, custom_fields=cf_payload or None
        )
        linked: list[str] = []
        defect_error = ""
        if defect_keys:
            try:
                self.xray.add_test_run_defects(run_id, defect_keys)
                linked = defect_keys
            except JiraError as exc:
                # Status already written — surface link failure without rolling back.
                defect_error = str(exc) or "Unable to link defects"
        self._bust_execution_case_cache(execution_key)
        result = {
            "ok": True,
            "run_id": str(run_id),
            "status": status,
            "defects": linked,
            "custom_fields_count": len(cf_payload),
        }
        if cf_skipped:
            result["custom_fields_skipped"] = cf_skipped
        if defect_error:
            result["defect_error"] = defect_error
            result["ok"] = False
        return result

    def add_case_defects(
        self,
        run_id: int | str,
        defects: Any = None,
        execution_key: str = "",
    ) -> dict[str, Any]:
        """Link existing Jira issue key(s) onto a Test Run (Execution Defects)."""
        defect_keys = self._normalize_defect_keys(defects)
        if not defect_keys:
            raise JiraError(
                "Provide at least one Jira issue key/ID (e.g. SI91X-12345), not a URL"
            )
        linked = self.xray.add_test_run_defects(run_id, defect_keys)
        self._bust_execution_case_cache(execution_key)
        return {
            "ok": True,
            "run_id": str(run_id),
            "defects": linked or defect_keys,
            "linked": defect_keys,
        }

    def add_case_defects_bulk(
        self,
        run_ids: list[int | str],
        defects: Any = None,
        execution_key: str = "",
    ) -> dict[str, Any]:
        """Link the same issue key(s) to many Test Runs (bulk Execution Defects)."""
        defect_keys = self._normalize_defect_keys(defects)
        if not defect_keys:
            raise JiraError(
                "Provide at least one Jira issue key/ID (e.g. SI91X-12345)"
            )
        updated: list[str] = []
        failed: list[dict[str, Any]] = []
        from concurrent.futures import ThreadPoolExecutor, as_completed

        def _one(run_id: int | str) -> tuple[str, Exception | None]:
            try:
                self.xray.add_test_run_defects(run_id, defect_keys)
                return str(run_id), None
            except Exception as exc:  # noqa: BLE001
                return str(run_id), exc

        with ThreadPoolExecutor(max_workers=6) as pool:
            futures = [pool.submit(_one, rid) for rid in run_ids if rid]
            for fut in as_completed(futures):
                run_id, err = fut.result()
                if err is None:
                    updated.append(run_id)
                else:
                    failed.append({"run_id": run_id, "error": str(err)})

        self._bust_execution_case_cache(execution_key)
        return {
            "ok": not failed,
            "defects": defect_keys,
            "updated": updated,
            "failed": failed,
            "updated_count": len(updated),
            "failed_count": len(failed),
        }

    def search_issues_for_defect_picker(
        self, query: str, *, limit: int = 20
    ) -> list[dict[str, Any]]:
        """Search Jira issues by key/text for the Execution Defects picker."""
        q = (query or "").strip()
        if not q:
            return []
        # Prefer issue picker (same UX source as Jira/Xray associate dialog).
        try:
            picker = self.jira.issue_picker(q, max_results=limit)
            if picker:
                return picker
        except JiraError:
            pass
        safe = self._jql_escape_phrase(q)
        if self._looks_like_issue_key(q):
            jql = f'key = "{q.strip().upper()}"'
        else:
            jql = f'summary ~ "{safe}" OR key = "{q.strip().upper()}" ORDER BY updated DESC'
        try:
            issues = self.jira.search(
                jql=jql,
                fields=["summary", "issuetype", "status"],
                max_results=limit,
            ).get("issues", [])
        except JiraError:
            return []
        rows: list[dict[str, Any]] = []
        for issue in issues or []:
            if not isinstance(issue, dict):
                continue
            key = (issue.get("key") or "").strip()
            if not key:
                continue
            fields = issue.get("fields") or {}
            itype = fields.get("issuetype") or {}
            status = fields.get("status") or {}
            rows.append(
                {
                    "key": key,
                    "summary": (fields.get("summary") or "").strip(),
                    "issuetype": (itype.get("name") or "").strip(),
                    "status": (status.get("name") or "").strip(),
                    "label": f"{key} — {(fields.get('summary') or '').strip()}".strip(
                        " —"
                    ),
                }
            )
        return rows

    def update_case_assignee(
        self, run_id: int | str, user: str, execution_key: str = ""
    ) -> dict[str, Any]:
        """Assign a Test Run to a Jira user (Xray Test Execution Assignee column)."""
        user = (user or "").strip()
        if not user:
            raise JiraError("user is required")
        # Allow pasting a display name → resolve to username when possible.
        user_key = user
        try:
            matches = self.jira.search_users(user, max_results=5)
        except JiraError:
            matches = []
        if matches:
            exact = next(
                (
                    m
                    for m in matches
                    if (m.get("name") or "").lower() == user.lower()
                    or (m.get("displayName") or "").lower() == user.lower()
                ),
                matches[0],
            )
            user_key = exact.get("name") or exact.get("key") or user
            display = exact.get("displayName") or user_key
        else:
            display = user
        self.xray.set_test_run_assignee(run_id, user_key)
        self._bust_execution_case_cache(execution_key)
        return {
            "ok": True,
            "run_id": str(run_id),
            "assignee": display,
            "assignee_key": user_key,
        }

    def update_case_assignees_bulk(
        self, run_ids: list[int | str], user: str, execution_key: str = ""
    ) -> dict[str, Any]:
        updated: list[str] = []
        failed: list[dict[str, Any]] = []
        from concurrent.futures import ThreadPoolExecutor, as_completed

        def _one(run_id: int | str) -> tuple[dict[str, Any], Exception | None]:
            try:
                return self.update_case_assignee(run_id, user, execution_key=""), None
            except Exception as exc:  # noqa: BLE001
                return {"run_id": str(run_id)}, exc

        with ThreadPoolExecutor(max_workers=6) as pool:
            futures = [pool.submit(_one, rid) for rid in run_ids if rid]
            for fut in as_completed(futures):
                payload, err = fut.result()
                if err is None:
                    updated.append(str(payload.get("run_id")))
                else:
                    failed.append({"run_id": payload.get("run_id"), "error": str(err)})

        self._bust_execution_case_cache(execution_key)
        return {
            "ok": not failed,
            "updated": updated,
            "failed": failed,
            "updated_count": len(updated),
            "failed_count": len(failed),
            "assignee": user,
        }

    def search_assignee_users(self, query: str) -> list[dict[str, Any]]:
        return self.jira.search_users(query, max_results=15)

    def update_case_statuses_bulk(
        self,
        run_ids: list[int | str],
        status: str,
        execution_key: str = "",
        defects: Any = None,
        custom_fields: Any = None,
    ) -> dict[str, Any]:
        status = self._normalize_status(status)
        if status not in EDITABLE_STATUSES and status not in STATUS_ALIASES.values():
            raise JiraError(f"Unsupported status: {status}")
        defect_keys = self._normalize_defect_keys(defects)
        cf_payload: list[dict[str, Any]] = []
        if isinstance(custom_fields, list):
            cf_payload = XrayClient._normalize_custom_fields(custom_fields)
        elif isinstance(custom_fields, dict) and custom_fields:
            catalog = (
                self.get_test_run_custom_field_catalog(execution_key).get("fields")
                if execution_key
                else empty_field_catalog()
            )
            cf_payload, _skipped = self.build_custom_field_payload(
                custom_fields, catalog
            )
        updated: list[str] = []
        failed: list[dict[str, Any]] = []
        from concurrent.futures import ThreadPoolExecutor, as_completed

        def _one(run_id: int | str) -> tuple[str, Exception | None]:
            try:
                self.xray.update_test_run(
                    run_id, status=status, custom_fields=cf_payload or None
                )
                if defect_keys:
                    self.xray.add_test_run_defects(run_id, defect_keys)
                return str(run_id), None
            except Exception as exc:  # noqa: BLE001
                return str(run_id), exc

        with ThreadPoolExecutor(max_workers=6) as pool:
            futures = [pool.submit(_one, rid) for rid in run_ids if rid]
            for fut in as_completed(futures):
                run_id, err = fut.result()
                if err is None:
                    updated.append(run_id)
                else:
                    failed.append({"run_id": run_id, "error": str(err)})

        self._bust_execution_case_cache(execution_key)

        return {
            "ok": not failed,
            "status": status,
            "updated": updated,
            "failed": failed,
            "updated_count": len(updated),
            "failed_count": len(failed),
            "defects": defect_keys,
        }

    @staticmethod
    def _failure_reason_text(
        *,
        explanation: str = "",
        key_error: str = "",
        test_name: str = "",
        summary: str = "",
    ) -> str:
        """Best available failure reason (KeyError/log), never the test path/title."""
        del test_name, summary  # intentionally unused — not failure reasons
        from .html_results import _clean_reason_text, extract_key_error_from_log

        key = _clean_reason_text(key_error or "")
        if key:
            return key[:2000]
        expl = (explanation or "").strip()
        if expl:
            short = extract_key_error_from_log(expl)
            return (short or _clean_reason_text(expl))[:2000]
        return ""

    @staticmethod
    def _failure_details_text(info: dict[str, Any]) -> str:
        """Compact extra failure info for UI/Excel (CLI API, not Python methods)."""
        bits: list[str] = []
        api = (info.get("failed_api") or info.get("failed_command") or "").strip()
        status = (info.get("status_code") or "").strip()
        ctx = (info.get("failure_context") or "").strip()
        if api:
            bits.append(f"API: {api}")
        if status:
            bits.append(f"Status: {status}")
        if ctx and "Call:" in ctx:
            call = ctx.split("Call:", 1)[-1].strip()
            if call:
                bits.append(f"Call: {call}")
        elif ctx and not bits:
            bits.append(ctx)
        return "\n".join(bits)[:800]

    @staticmethod
    def _jql_escape_phrase(value: str) -> str:
        return (value or "").replace("\\", "\\\\").replace('"', '\\"')

    @classmethod
    def _search_phrase_from_reason(cls, reason: str) -> str:
        """Turn a failure reason into a short JQL text~ phrase."""
        import re

        text = (reason or "").strip()
        if not text:
            return ""
        first = text.splitlines()[0].strip()
        first = re.sub(
            r"^(FAILED|ERROR|FAILURE|AssertionError|KeyError|Exception|Error)\s*:?\s*",
            "",
            first,
            flags=re.I,
        )
        cleaned = re.sub(r"[+\-&|!(){}\[\]^~*?:\\/\"]+", " ", first)
        stop = {
            "the",
            "and",
            "for",
            "with",
            "from",
            "this",
            "that",
            "none",
            "null",
            "true",
            "false",
            "test",
            "failed",
            "failure",
            "error",
            "assert",
            "self",
        }
        tokens = [
            t
            for t in cleaned.split()
            if len(t) >= 3 and not t.isdigit() and t.lower() not in stop
        ]
        if not tokens:
            return ""
        phrase = " ".join(tokens[:8])
        if len(phrase) > 120:
            phrase = phrase[:120].rsplit(" ", 1)[0]
        return phrase.strip()

    @staticmethod
    def _strip_jql_order_by(jql: str) -> str:
        import re

        return re.sub(
            r"\s+order\s+by\s+.+$", "", (jql or "").strip(), flags=re.I
        ).strip()

    def _similar_bug_base_jql(self) -> str:
        raw = (getattr(settings, "JIRA_SIMILAR_BUG_BASE_JQL", "") or "").strip()
        return self._strip_jql_order_by(raw)

    def _similar_bug_extra_jqls(self) -> list[str]:
        raw = (getattr(settings, "JIRA_SIMILAR_BUG_EXTRA_JQL", "") or "").strip()
        if not raw:
            return []
        parts = [self._strip_jql_order_by(p) for p in raw.split(";")]
        return [p for p in parts if p]

    def _pack_similar_bug_issue(
        self, issue: dict[str, Any], *, pool: str
    ) -> dict[str, Any] | None:
        key = (issue.get("key") or "").strip()
        if not key:
            return None
        f = issue.get("fields") or {}
        summary = f.get("summary") or ""
        description = self._plain_description(f.get("description"))
        hay = f"{summary}\n{description}".lower()
        return {
            "key": key,
            "summary": summary,
            "status": ((f.get("status") or {}).get("name") or ""),
            "priority": ((f.get("priority") or {}).get("name") or ""),
            "issue_type": ((f.get("issuetype") or {}).get("name") or ""),
            "assignee": self._user_name(f.get("assignee")),
            "reporter": self._user_name(f.get("reporter")),
            "haystack": hay,
            "created": f.get("created"),
            "updated": f.get("updated"),
            "pool": pool,
        }

    def _fetch_similar_bug_pool(
        self, jql: str, *, pool: str, hard_limit: int
    ) -> list[dict[str, Any]]:
        fields = [
            "summary",
            "status",
            "priority",
            "issuetype",
            "description",
            "assignee",
            "reporter",
            "created",
            "updated",
        ]
        try:
            issues = self.jira.search_all(
                jql=f"{jql} ORDER BY created DESC",
                fields=fields,
                page_size=100,
                hard_limit=hard_limit,
            )
        except JiraError:
            return []
        out: list[dict[str, Any]] = []
        for issue in issues or []:
            packed = self._pack_similar_bug_issue(issue, pool=pool)
            if packed:
                out.append(packed)
        return out

    def _load_similar_bug_corpus(self) -> list[dict[str, Any]]:
        """Load Coex (preferred) + SI91X/extra Bug pools; cache for FAIL matching."""
        base = self._similar_bug_base_jql()
        extras = self._similar_bug_extra_jqls()
        if not base and not extras:
            return []
        primary_limit = int(
            getattr(settings, "JIRA_SIMILAR_BUG_CORPUS_LIMIT", 500) or 500
        )
        primary_limit = max(50, min(primary_limit, 2000))
        extra_limit = int(
            getattr(settings, "JIRA_SIMILAR_BUG_EXTRA_CORPUS_LIMIT", 300) or 300
        )
        extra_limit = max(50, min(extra_limit, 2000))
        cache_key = (
            "similar_bug_corpus_v3:"
            f"{base}|{';'.join(extras)}|{primary_limit}|{extra_limit}"
        )
        cached = cache.get(cache_key)
        if isinstance(cached, list) and cached:
            return cached

        corpus: list[dict[str, Any]] = []
        seen: set[str] = set()
        if base:
            for bug in self._fetch_similar_bug_pool(
                base, pool="coex", hard_limit=primary_limit
            ):
                key = bug["key"]
                if key in seen:
                    continue
                seen.add(key)
                corpus.append(bug)
        for idx, extra in enumerate(extras):
            pool_name = "si91x" if "SI91X" in extra.upper() else f"extra{idx}"
            for bug in self._fetch_similar_bug_pool(
                extra, pool=pool_name, hard_limit=extra_limit
            ):
                key = bug["key"]
                if key in seen:
                    continue
                seen.add(key)
                corpus.append(bug)
        if corpus:
            cache.set(cache_key, corpus, max(settings.JIRA_CACHE_SECONDS, 600))
        return corpus

    @classmethod
    def _score_bug_for_phrase(cls, bug: dict[str, Any], phrase: str) -> int:
        """Token-overlap score of a corpus bug against a failure search phrase."""
        tokens = [t for t in (phrase or "").lower().split() if len(t) >= 3]
        if not tokens:
            return 0
        hay = bug.get("haystack") or ""
        if not hay:
            hay = f"{bug.get('summary') or ''}".lower()
        score = 0
        for tok in tokens:
            if tok in hay:
                score += 2 if len(tok) >= 6 else 1
        # Bonus when several phrase tokens appear as a cluster in the summary.
        summary = (bug.get("summary") or "").lower()
        joined = " ".join(tokens[:4])
        if joined and joined in summary:
            score += 4
        # Prefer Coex SQA pool when scores are otherwise close.
        if (bug.get("pool") or "") == "coex":
            score += 1
        return score

    def _rank_similar_from_corpus(
        self,
        phrase: str,
        *,
        corpus: list[dict[str, Any]],
        exclude_keys: set[str] | None = None,
        limit: int = 5,
    ) -> list[dict[str, Any]]:
        exclude = {k for k in (exclude_keys or set()) if k}
        scored: list[tuple[int, int, dict[str, Any]]] = []
        for bug in corpus:
            key = bug.get("key") or ""
            if not key or key in exclude:
                continue
            score = self._score_bug_for_phrase(bug, phrase)
            if score <= 0:
                continue
            # Primary pool preferred on ties (coex=1, others=0).
            pool_rank = 1 if (bug.get("pool") or "") == "coex" else 0
            scored.append((score, pool_rank, bug))
        scored.sort(
            key=lambda pair: (pair[0], pair[1], pair[2].get("created") or ""),
            reverse=True,
        )

        results: list[dict[str, Any]] = []
        for score, _pool_rank, bug in scored[:limit]:
            results.append(
                {
                    "key": bug.get("key") or "",
                    "summary": bug.get("summary") or "",
                    "status": bug.get("status") or "",
                    "priority": bug.get("priority") or "",
                    "issue_type": bug.get("issue_type") or "",
                    "assignee": bug.get("assignee") or "",
                    "reporter": bug.get("reporter") or "",
                    "pool": bug.get("pool") or "",
                    "match_query": phrase,
                    "match_score": score,
                    "search_jql": self._similar_bug_base_jql(),
                    "url": self.jira.browse_url(bug.get("key") or ""),
                }
            )
        return results

    def _search_similar_jiras(
        self,
        reason: str,
        *,
        exclude_keys: set[str] | None = None,
        limit: int = 5,
        corpus: list[dict[str, Any]] | None = None,
    ) -> list[dict[str, Any]]:
        """Read-only similar-bug search: Coex SQA preferred, SI91X also surfed."""
        phrase = self._search_phrase_from_reason(reason)
        if not phrase:
            # reason may already be a cleaned phrase from _attach_similar_jiras
            phrase = (reason or "").strip()
        if not phrase:
            return []

        base = self._similar_bug_base_jql()
        extras = self._similar_bug_extra_jqls()
        if base or extras:
            pool = corpus if corpus is not None else self._load_similar_bug_corpus()
            if pool:
                return self._rank_similar_from_corpus(
                    phrase,
                    corpus=pool,
                    exclude_keys=exclude_keys,
                    limit=limit,
                )
            # Corpus empty (JQL failed) — try text search inside the primary filter.
            if base:
                return self._search_similar_jiras_jql(
                    phrase,
                    base_jql=base,
                    exclude_keys=exclude_keys,
                    limit=limit,
                )

        # Legacy project-key search when no base/extra JQL configured.
        return self._search_similar_jiras_jql(
            phrase,
            base_jql="",
            exclude_keys=exclude_keys,
            limit=limit,
        )

    def _search_similar_jiras_jql(
        self,
        phrase: str,
        *,
        base_jql: str = "",
        exclude_keys: set[str] | None = None,
        limit: int = 5,
    ) -> list[dict[str, Any]]:
        """JQL text~ search, optionally constrained by base_jql / project keys."""
        exclude = {k for k in (exclude_keys or set()) if k}
        safe = self._jql_escape_phrase(phrase)
        if not safe:
            return []

        jql_attempts: list[str] = []
        base = self._strip_jql_order_by(base_jql)
        if base:
            jql_attempts.append(
                f"{base} AND "
                f'(summary ~ "{safe}" OR description ~ "{safe}" OR text ~ "{safe}") '
                "ORDER BY created DESC"
            )
            jql_attempts.append(
                f'{base} AND (summary ~ "{safe}" OR text ~ "{safe}") '
                "ORDER BY created DESC"
            )
        else:
            bug_types = [
                t.strip()
                for t in (getattr(settings, "JIRA_BUG_ISSUE_TYPES", "") or "").split(",")
                if t.strip()
            ]
            projects = [
                p.strip()
                for p in (getattr(settings, "JIRA_BUG_PROJECT_KEYS", "") or "").split(",")
                if p.strip()
            ]
            type_clause = ""
            if bug_types:
                type_clause = (
                    "issuetype in ("
                    + ", ".join(f'"{t}"' for t in bug_types)
                    + ") AND "
                )
            proj_clause = ""
            if projects:
                proj_clause = (
                    "project in (" + ", ".join(f'"{p}"' for p in projects) + ") AND "
                )
            if proj_clause and type_clause:
                jql_attempts.append(
                    f"{proj_clause}{type_clause}"
                    f'(summary ~ "{safe}" OR description ~ "{safe}" OR text ~ "{safe}") '
                    "ORDER BY created DESC"
                )
            elif proj_clause:
                jql_attempts.append(
                    f"{proj_clause}"
                    f'(summary ~ "{safe}" OR text ~ "{safe}") ORDER BY created DESC'
                )

        cache_key = f"similar_jiras_v2:{phrase}|{base}|{limit}"
        cached = cache.get(cache_key)
        if isinstance(cached, list):
            results = [
                {**item, "url": self.jira.browse_url(item.get("key") or "")}
                for item in cached
                if item.get("key") and item.get("key") not in exclude
            ]
            return results[:limit]

        fields = ["summary", "status", "priority", "issuetype", "updated", "created"]
        issues: list[dict[str, Any]] = []
        used_jql = ""
        for jql in jql_attempts:
            try:
                issues = self.jira.search(
                    jql=jql, fields=fields, max_results=max(limit * 3, 15)
                ).get("issues", [])
            except JiraError:
                continue
            if issues:
                used_jql = jql
                break

        raw_results: list[dict[str, Any]] = []
        for issue in issues or []:
            key = (issue.get("key") or "").strip()
            if not key:
                continue
            f = issue.get("fields") or {}
            raw_results.append(
                {
                    "key": key,
                    "summary": f.get("summary") or "",
                    "status": ((f.get("status") or {}).get("name") or ""),
                    "priority": ((f.get("priority") or {}).get("name") or ""),
                    "issue_type": ((f.get("issuetype") or {}).get("name") or ""),
                    "match_query": phrase,
                    "search_jql": used_jql,
                }
            )
            if len(raw_results) >= max(limit * 2, 10):
                break
        if raw_results:
            cache.set(cache_key, raw_results, max(settings.JIRA_CACHE_SECONDS, 300))

        results = [
            {**item, "url": self.jira.browse_url(item.get("key") or "")}
            for item in raw_results
            if item.get("key") not in exclude
        ]
        return results[:limit]

    def _attach_similar_jiras(self, rows: list[dict[str, Any]]) -> None:
        """Group rows by search phrase and attach similar_jiras (read-only)."""
        from concurrent.futures import ThreadPoolExecutor, as_completed

        by_phrase: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            seed = " ".join(
                part
                for part in (
                    row.get("reason") or "",
                    row.get("failed_api") or "",
                    row.get("failed_command") or "",
                    row.get("status_code") or "",
                )
                if part
            )
            phrase = self._search_phrase_from_reason(seed)
            row["search_phrase"] = phrase
            if phrase:
                by_phrase[phrase].append(row)
            else:
                row["similar_jiras"] = []
                row["similar_jiras_text"] = ""

        # Cap unique searches so large FAIL sets stay responsive.
        phrase_list = list(by_phrase.keys())[:80]
        found: dict[str, list[dict[str, Any]]] = {}

        # Load Coex (preferred) + SI91X/extra pools once, then rank per failure phrase.
        corpus = (
            self._load_similar_bug_corpus()
            if (self._similar_bug_base_jql() or self._similar_bug_extra_jqls())
            else []
        )

        def _one(phrase: str) -> tuple[str, list[dict[str, Any]]]:
            linked: set[str] = set()
            for row in by_phrase.get(phrase) or []:
                for k in row.get("existing_defects") or []:
                    if k:
                        linked.add(str(k))
            if corpus:
                return phrase, self._rank_similar_from_corpus(
                    phrase, corpus=corpus, exclude_keys=linked, limit=5
                )
            return phrase, self._search_similar_jiras(
                phrase, exclude_keys=linked, limit=5, corpus=corpus or None
            )

        if phrase_list:
            # Local ranking is cheap; keep a small pool for JQL fallback paths.
            workers = 1 if corpus else min(6, len(phrase_list))
            with ThreadPoolExecutor(max_workers=workers) as pool:
                futs = [pool.submit(_one, p) for p in phrase_list]
                for fut in as_completed(futs):
                    try:
                        phrase, hits = fut.result()
                    except Exception:  # noqa: BLE001
                        continue
                    found[phrase] = hits

        for phrase, group in by_phrase.items():
            hits = found.get(phrase) or []
            text = self._cell_list(
                [
                    f"{h.get('key')}: {(h.get('summary') or '')[:120]}".strip()
                    for h in hits
                    if h.get("key")
                ]
            )
            for row in group:
                row["similar_jiras"] = hits
                row["similar_jiras_text"] = text

    def _build_failure_triage(
        self,
        *,
        execution_key: str,
        statuses: dict[str, str],
        by_map: dict[str, list[dict[str, Any]]],
        report_meta: dict[str, dict[str, str]] | None = None,
        execution_cases: list[dict[str, Any]] | None = None,
    ) -> dict[str, Any]:
        """Preview-only triage: Xray TODO / untested cases only.

        One row per untested case. Report FAIL/ERROR enriches reason/API when
        present. Already PASS (or any non-TODO) in Xray is out of scope.
        Does not write to Jira/Xray.
        """
        meta = report_meta or {}
        failures: list[dict[str, Any]] = []
        unmatched_failures: list[dict[str, Any]] = []
        todo_cases: list[dict[str, Any]] = []
        skipped_xray_pass = 0

        def _is_report_fail(status: str) -> bool:
            return self._normalize_status(status) == "FAIL"

        # Count report FAILs that map to already-PASS cases (not triaged).
        for map_id, st in statuses.items():
            if not _is_report_fail(st):
                continue
            for case in by_map.get(map_id) or []:
                if self._normalize_status(case.get("status") or "TODO") == "PASS":
                    skipped_xray_pass += 1

        todo_total = 0
        for case in execution_cases or []:
            current = self._normalize_status(case.get("status") or "TODO")
            if current != "TODO":
                continue
            todo_total += 1
            map_id = normalize_map_id(case.get("test_src_map_id") or "")
            report_status = self._normalize_status(statuses.get(map_id) or "")
            info = meta.get(map_id) or {}
            has_report_fail = bool(map_id and _is_report_fail(report_status))
            if has_report_fail:
                base_extra = {
                    "test_name": info.get("test_name") or "",
                    "technology": info.get("technology") or "",
                    "explanation": info.get("explanation") or "",
                    "key_error": info.get("key_error") or "",
                    "source": info.get("source") or "",
                    "failed_api": info.get("failed_api") or "",
                    "failed_command": info.get("failed_command") or "",
                    "status_code": info.get("status_code") or "",
                    "failure_context": info.get("failure_context") or "",
                }
                reason = self._failure_reason_text(
                    explanation=base_extra["explanation"],
                    key_error=base_extra["key_error"],
                    test_name=base_extra["test_name"],
                    summary=case.get("summary") or "",
                )
                html_status = "FAIL"
            else:
                base_extra = {
                    "test_name": "",
                    "technology": case.get("technology") or "",
                    "explanation": "",
                    "key_error": "",
                    "source": "",
                    "failed_api": "",
                    "failed_command": "",
                    "status_code": "",
                    "failure_context": "",
                }
                reason = "Untested (TODO in Xray)"
                html_status = ""
            api = base_extra["failed_api"] or base_extra["failed_command"]
            section_path = self._clean_path(case.get("section_path") or "")
            section_name = section_path.split("/")[-1] if section_path else ""
            defect_keys = self._defect_keys(case.get("defects"))
            failures.append(
                {
                    "test_src_map_id": map_id,
                    "html_status": html_status,
                    "key": (case.get("key") or "").strip(),
                    "summary": case.get("summary") or "",
                    "run_id": case.get("run_id") or "",
                    "current_status": "TODO",
                    "url": case.get("url") or "",
                    "matched": True,
                    "triage_kind": "fail",
                    "untested_fail": not has_report_fail,
                    "existing_defects": defect_keys,
                    "section_path": section_path,
                    "section": section_name,
                    "technology": base_extra["technology"]
                    or (case.get("technology") or ""),
                    "test_name": base_extra["test_name"],
                    "explanation": base_extra["explanation"],
                    "key_error": base_extra["key_error"],
                    "source": base_extra["source"],
                    "failed_api": api,
                    "failed_command": base_extra["failed_command"] or api,
                    "status_code": base_extra["status_code"],
                    "failure_context": base_extra["failure_context"],
                    "reason": reason,
                    "reason_short": (reason.splitlines()[0][:280] if reason else ""),
                    "details_text": self._failure_details_text(
                        {**base_extra, "failed_api": api}
                    )
                    if has_report_fail
                    else "",
                }
            )

        # Enrich linked defects with Jira status/priority (batched, read-only).
        all_defect_keys: list[str] = []
        seen: set[str] = set()
        for row in failures:
            for dkey in row.get("existing_defects") or []:
                if dkey not in seen:
                    seen.add(dkey)
                    all_defect_keys.append(dkey)
        defect_details = (
            self._fetch_defect_issue_fields(all_defect_keys) if all_defect_keys else {}
        )
        for row in failures:
            keys = row.get("existing_defects") or []
            status_bits: list[str] = []
            priority_bits: list[str] = []
            for dkey in keys:
                dinfo = defect_details.get(dkey) or {}
                status_bits.append((dinfo.get("status") or "").strip())
                priority_bits.append((dinfo.get("priority") or "").strip())
            row["defect_statuses"] = status_bits
            row["defect_priorities"] = priority_bits
            row["existing_defects_text"] = self._cell_list(keys)
            row["defect_statuses_text"] = self._cell_list(
                status_bits, keep_blanks=True
            )
            row["defect_priorities_text"] = self._cell_list(
                priority_bits, keep_blanks=True
            )

        # Similar bugs only when the upload supplied failure text.
        self._attach_similar_jiras(
            [row for row in failures if (row.get("html_status") or "") == "FAIL"]
        )

        report_status_counts = {
            status: sum(1 for s in statuses.values() if s == status)
            for status in sorted({self._normalize_status(s) for s in statuses.values()})
        }
        return {
            "execution": execution_key,
            "failures": failures,
            "failure_count": len(failures),
            "todo_cases": todo_cases,
            # Full execution TODO/untested count (matches dashboard).
            "todo_count": todo_total,
            "unmatched_failures": unmatched_failures,
            "unmatched_failure_count": len(unmatched_failures),
            "skipped_xray_pass_count": skipped_xray_pass,
            "report_fail_on_xray_pass_count": 0,
            "triage_status_counts": report_status_counts,
            "triage_mode": "preview_only",
            "writes_jira": False,
            "writes_xray": False,
        }

    def build_failure_triage_xlsx(
        self,
        *,
        execution_key: str,
        failures: list[dict[str, Any]],
        unmatched_failures: list[dict[str, Any]] | None = None,
        todo_cases: list[dict[str, Any]] | None = None,
    ) -> bytes:
        """Build a downloadable Failure triage workbook from preview rows (read-only)."""
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Failure triage"
        ws.append(
            [
                "Test Execution",
                "Kind",
                "test_src_map_id",
                "Case",
                "Title",
                "Section",
                "Section path",
                "Report status",
                "Current Xray status",
                "Reason",
                "Failed API",
                "Failed command",
                "Status code",
                "Failure context",
                "Search phrase",
                "Technology",
                "Test name",
                "Explanation",
                "Key error",
                "Source",
                "Run ID",
                "Linked Jiras (on case)",
                "Selected Jira",
                "Selected status",
                "Selected priority",
                "Selected title",
                "Selected reporter",
                "Selected assignee",
                "Similar Jiras (search)",
                "Matched",
                "Case URL",
            ]
        )

        from .html_results import _clean_reason_text

        def _keep_row(row: dict[str, Any]) -> bool:
            # Drop rows already PASS in Xray; keep FAIL / untested triage rows.
            current = self._normalize_status(row.get("current_status") or "")
            if row.get("matched") is False:
                return True
            return current != "PASS"

        def _row_kind(row: dict[str, Any]) -> str:
            kind = (row.get("triage_kind") or "").strip().lower()
            if kind == "todo":
                return "TODO"
            if kind == "fail":
                return "FAIL"
            if self._normalize_status(row.get("current_status") or "") == "TODO" and not (
                row.get("html_status") or ""
            ):
                return "TODO"
            return "FAIL"

        def _issue_lookup(key: str) -> dict[str, Any]:
            k = (key or "").strip()
            if not k:
                return {}
            if k in issue_fields:
                return issue_fields[k]
            upper = k.upper()
            for ik, info in issue_fields.items():
                if (ik or "").upper() == upper:
                    return info
            return {}

        rows = [r for r in (failures or []) if _keep_row(r)]
        for row in unmatched_failures or []:
            if _keep_row({**row, "matched": False}):
                rows.append({**row, "matched": False, "triage_kind": "fail"})
        for row in todo_cases or []:
            if _keep_row(row):
                rows.append({**row, "triage_kind": row.get("triage_kind") or "todo"})

        # Fetch title/status/people for selected tickets (and any still-linked keys).
        keys_to_fetch: list[str] = []
        seen_keys: set[str] = set()
        for row in rows:
            for key in (
                (row.get("selected_jira") or "").strip(),
                *(
                    row.get("existing_defects")
                    if isinstance(row.get("existing_defects"), list)
                    else []
                ),
            ):
                k = str(key or "").strip()
                if k and self._looks_like_issue_key(k) and k not in seen_keys:
                    seen_keys.add(k)
                    keys_to_fetch.append(k)
        issue_fields = (
            self._fetch_defect_issue_fields(keys_to_fetch) if keys_to_fetch else {}
        )

        # Aggregate selected Jiras for the summary tab.
        jira_summary: dict[str, dict[str, Any]] = {}

        for row in rows:
            defects = row.get("existing_defects") or []
            if isinstance(defects, str):
                linked_text = defects
            else:
                linked_text = row.get("existing_defects_text") or self._cell_list(
                    list(defects)
                )
            similar = row.get("similar_jiras_text") or ""
            if not similar:
                similar_list = row.get("similar_jiras") or []
                if isinstance(similar_list, list):
                    similar = self._cell_list(
                        [
                            f"{h.get('key')}: {(h.get('summary') or '')[:120]}".strip()
                            for h in similar_list
                            if isinstance(h, dict) and h.get("key")
                        ]
                    )
            reason = _clean_reason_text(
                row.get("reason")
                or self._failure_reason_text(
                    explanation=row.get("explanation") or "",
                    key_error=row.get("key_error") or "",
                    test_name=row.get("test_name") or "",
                    summary=row.get("summary") or "",
                )
            )
            api = (row.get("failed_api") or row.get("failed_command") or "").strip()
            selected = (row.get("selected_jira") or "").strip()
            selected_info = dict(_issue_lookup(selected))
            # Fall back to similar_jiras payload from preview when fetch misses.
            if selected and not (
                selected_info.get("status") or selected_info.get("summary")
            ):
                for hit in row.get("similar_jiras") or []:
                    if isinstance(hit, dict) and (hit.get("key") or "") == selected:
                        selected_info = {
                            "summary": hit.get("summary") or "",
                            "status": hit.get("status") or "",
                            "priority": hit.get("priority") or "",
                            "assignee": hit.get("assignee") or "",
                            "reporter": hit.get("reporter") or "",
                        }
                        break
            kind = _row_kind(row)
            if selected and self._looks_like_issue_key(selected):
                canon = selected.upper()
                entry = jira_summary.setdefault(
                    canon,
                    {
                        "key": selected,
                        "fail_count": 0,
                        "todo_count": 0,
                        "total_count": 0,
                        "case_keys": [],
                        "map_ids": [],
                        "info": selected_info,
                    },
                )
                entry["total_count"] += 1
                if kind == "TODO":
                    entry["todo_count"] += 1
                else:
                    entry["fail_count"] += 1
                case_key = (row.get("key") or "").strip()
                map_id = (row.get("test_src_map_id") or "").strip()
                if case_key and case_key not in entry["case_keys"]:
                    entry["case_keys"].append(case_key)
                if map_id and map_id not in entry["map_ids"]:
                    entry["map_ids"].append(map_id)
                # Prefer fuller field payload when available.
                if selected_info and (
                    not entry["info"]
                    or (selected_info.get("summary") and not entry["info"].get("summary"))
                ):
                    entry["info"] = selected_info
            section_path = self._clean_path(row.get("section_path") or "")
            section_name = (row.get("section") or "").strip() or (
                section_path.split("/")[-1] if section_path else ""
            )
            ws.append(
                [
                    execution_key or "",
                    kind,
                    row.get("test_src_map_id") or "",
                    row.get("key") or "",
                    row.get("summary") or "",
                    section_name,
                    section_path,
                    row.get("html_status")
                    or ("TODO" if kind == "TODO" else "FAIL"),
                    row.get("current_status") or "",
                    reason,
                    api,
                    row.get("failed_command") or api,
                    row.get("status_code") or "",
                    row.get("failure_context") or "",
                    row.get("search_phrase") or "",
                    row.get("technology") or "",
                    row.get("test_name") or "",
                    row.get("explanation") or "",
                    _clean_reason_text(row.get("key_error") or ""),
                    row.get("source") or "",
                    row.get("run_id") or "",
                    linked_text,
                    selected,
                    (selected_info.get("status") or "").strip(),
                    (selected_info.get("priority") or "").strip(),
                    (selected_info.get("summary") or "").strip(),
                    (selected_info.get("reporter") or "").strip(),
                    (selected_info.get("assignee") or "").strip(),
                    similar,
                    "Yes" if row.get("matched") else "No",
                    row.get("url") or "",
                ]
            )

        self._autosize_sheet(ws)

        # Summary tab: matches the team defect-impact table layout.
        summary_ws = wb.create_sheet("Jira summary")
        summary_ws.append(
            [
                "Type",
                "Issue key",
                "No of TCs Impacted",
                "Summary",
                "Status",
                "Priority",
                "Assignee",
            ]
        )
        total_impacted = 0
        for entry in sorted(
            jira_summary.values(),
            key=lambda e: (
                -int(e.get("total_count") or 0),
                -int(e.get("fail_count") or 0),
                e.get("key") or "",
            ),
        ):
            info = entry.get("info") or _issue_lookup(entry.get("key") or "")
            jira_key = entry.get("key") or ""
            # Prefer unique case keys; fall back to row count.
            impacted = len(entry.get("case_keys") or []) or int(
                entry.get("total_count") or 0
            )
            total_impacted += impacted
            summary_ws.append(
                [
                    (info.get("issuetype") or "Bug").strip() or "Bug",
                    jira_key,
                    impacted,
                    (info.get("summary") or "").strip(),
                    (info.get("status") or "").strip(),
                    (info.get("priority") or "").strip(),
                    (info.get("assignee") or "").strip(),
                ]
            )
        summary_ws.append(
            [
                "Total",
                "",
                total_impacted,
                "",
                "",
                "",
                "",
            ]
        )
        self._autosize_sheet(summary_ws)

        todo_rows = [r for r in rows if _row_kind(r) == "TODO"]
        fail_rows = [r for r in rows if _row_kind(r) != "TODO"]
        meta = wb.create_sheet("Export info", 0)
        meta.append(["Scope", "Failure triage preview (read-only)"])
        meta.append(["Writes to Jira/Xray", "No"])
        meta.append(
            [
                "Filter",
                "Xray TODO/untested only (report FAIL details when present; PASS skipped)",
            ]
        )
        meta.append(["Test Execution", execution_key or ""])
        meta.append(["Open FAIL rows", len(fail_rows)])
        meta.append(["TODO / untested rows", len(todo_rows)])
        meta.append(
            ["Unmatched failures", len([r for r in rows if not r.get("matched")])]
        )
        meta.append(["Selected Jiras (summary)", len(jira_summary)])
        self._autosize_sheet(meta)

        from io import BytesIO

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def preview_html_import(
        self,
        execution_key: str,
        *,
        folder_paths: list[str] | None = None,
        uploaded_files: list[tuple[str, bytes]] | None = None,
        technology: str | None = None,
        only_changed: bool = True,
    ) -> dict[str, Any]:
        """Parse HTML reports; only propose PASS updates. FAIL and already-PASS are skipped.

        Multiple files/folders are processed in one preview. Duplicate files
        (same basename or content hash) are dropped; later filename timestamp
        wins for the same C###### across reports.
        """
        if not execution_key:
            raise JiraError("execution is required")

        # Match against all cases in the execution (ignore UI technology filter).
        match_technology = ""
        blobs: list[tuple[str, bytes]] = []
        path_errors: list[ParsedHtmlReport] = []
        for name, data in uploaded_files or []:
            blobs.append((name or "upload", data or b""))
        for location in folder_paths or []:
            loc = (location or "").strip()
            if not loc:
                continue
            try:
                paths = collect_html_paths(loc, recursive=True)
            except (FileNotFoundError, ValueError, OSError) as exc:
                bad = ParsedHtmlReport(
                    source=loc, timestamp=extract_unix_timestamp(loc)
                )
                bad.errors.append(str(exc))
                path_errors.append(bad)
                continue
            if not paths:
                empty = ParsedHtmlReport(
                    source=loc, timestamp=extract_unix_timestamp(loc)
                )
                empty.errors.append("No .htm/.html files found")
                path_errors.append(empty)
                continue
            for path in paths:
                try:
                    blobs.append((str(path), path.read_bytes()))
                except OSError as exc:
                    bad = ParsedHtmlReport(
                        source=str(path),
                        timestamp=extract_unix_timestamp(str(path)),
                    )
                    bad.errors.append(f"Unable to read {path}: {exc}")
                    path_errors.append(bad)

        # Content-hash only so same basename FAIL/PASS reports both stay for triage.
        blobs, skipped_dupes = dedupe_named_blobs(blobs, by_basename=False)
        reports = [parse_html_bytes(data, source=name) for name, data in blobs]
        reports.extend(path_errors)

        if not reports:
            raise JiraError("Provide a folder/file path and/or upload HTML files")

        # PASS apply: later timestamp wins. Triage: worst status wins so FAIL
        # still appears when a later report in the same upload says PASS.
        html_statuses, _chrono_meta = merge_reports_chronological_detailed(reports)
        triage_statuses, triage_meta = merge_reports_detailed(reports)
        cases = self.load_execution_cases(
            execution_key, technology=match_technology, force_refresh=True
        )

        by_map: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for case in cases:
            map_id = normalize_map_id(case.get("test_src_map_id") or "")
            if map_id:
                by_map[map_id].append(case)

        matches: list[dict[str, Any]] = []
        skipped: list[dict[str, Any]] = []
        unmatched_html: list[dict[str, Any]] = []
        no_run_id: list[dict[str, Any]] = []

        for map_id, new_status in sorted(html_statuses.items()):
            case_list = by_map.get(map_id) or []
            if not case_list:
                unmatched_html.append(
                    {"test_src_map_id": map_id, "html_status": new_status}
                )
                continue
            for case in case_list:
                current = self._normalize_status(case.get("status") or "TODO")
                changed = current != new_status
                item = {
                    "test_src_map_id": map_id,
                    "key": case.get("key"),
                    "summary": case.get("summary") or "",
                    "run_id": case.get("run_id"),
                    "current_status": current,
                    "html_status": new_status,
                    "changed": changed,
                    "url": case.get("url") or "",
                    "will_update": new_status == "PASS" and current != "PASS",
                }
                if not case.get("run_id"):
                    no_run_id.append(item)
                    continue
                # From HTML: never write FAIL/TODO/etc. — only PASS.
                if new_status != "PASS":
                    skipped.append({**item, "skip_reason": "non_pass"})
                    continue
                # Already PASS in Xray — leave untouched.
                if current == "PASS":
                    skipped.append({**item, "skip_reason": "already_present"})
                    continue
                if only_changed and not changed:
                    skipped.append({**item, "skip_reason": "unchanged"})
                    continue
                matches.append(item)

        parse_errors = []
        for report in reports:
            parse_errors.extend(
                {"source": report.source, "error": err} for err in report.errors
            )

        ordered_files = sorted(
            reports,
            key=lambda r: (r.timestamp if r.timestamp is not None else -1, r.source or ""),
        )

        triage = self._build_failure_triage(
            execution_key=execution_key,
            statuses=triage_statuses,
            by_map=by_map,
            report_meta=triage_meta,
            execution_cases=cases,
        )

        return {
            "ok": True,
            "execution": execution_key,
            "technology": technology or "",
            "mode": "html_pass_only",
            "matched_against": "all_execution_cases",
            "files": [r.to_dict() for r in ordered_files],
            "parsed_map_ids": len(html_statuses),
            "matches": matches,
            "match_count": len(matches),
            "skipped": skipped,
            "skipped_count": len(skipped),
            "unmatched_html": unmatched_html,
            "unmatched_count": len(unmatched_html),
            "no_run_id": no_run_id,
            "no_run_id_count": len(no_run_id),
            "parse_errors": parse_errors,
            "duplicate_files_skipped": skipped_dupes,
            "duplicate_files_skipped_count": len(skipped_dupes),
            "status_counts": {
                status: sum(1 for s in html_statuses.values() if s == status)
                for status in sorted(set(html_statuses.values()))
            },
            **triage,
        }

    def preview_excel_pass_import(
        self,
        execution_key: str,
        *,
        excel_bytes: bytes,
        excel_name: str = "upload.xlsx",
    ) -> dict[str, Any]:
        """Parse OverAllStatus-style Excel; later Time wins; preview PASS-only updates."""
        if not execution_key:
            raise JiraError("execution is required")
        if not excel_bytes:
            raise JiraError("Excel file is required")

        report = parse_excel_bytes(excel_bytes, source=excel_name or "upload.xlsx")
        if report.errors and not report.by_map_id:
            raise JiraError("; ".join(report.errors))

        excel_statuses = report.by_map_id
        winning_sources = report.winning_sources
        # PASS apply uses later Time; triage prefers FAIL if any row failed.
        triage_statuses, triage_meta = excel_triage_statuses(report)
        cases = self.load_execution_cases(
            execution_key, technology="", force_refresh=True
        )

        by_map: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for case in cases:
            map_id = normalize_map_id(case.get("test_src_map_id") or "")
            if map_id:
                by_map[map_id].append(case)

        matches: list[dict[str, Any]] = []
        skipped: list[dict[str, Any]] = []
        unmatched_html: list[dict[str, Any]] = []
        no_run_id: list[dict[str, Any]] = []

        for map_id, new_status in sorted(excel_statuses.items()):
            case_list = by_map.get(map_id) or []
            source = winning_sources.get(map_id) or report.source
            if not case_list:
                unmatched_html.append(
                    {
                        "test_src_map_id": map_id,
                        "html_status": new_status,
                        "source": source,
                    }
                )
                continue
            for case in case_list:
                current = self._normalize_status(case.get("status") or "TODO")
                changed = current != new_status
                item = {
                    "test_src_map_id": map_id,
                    "key": case.get("key"),
                    "summary": case.get("summary") or "",
                    "run_id": case.get("run_id"),
                    "current_status": current,
                    "html_status": new_status,
                    "changed": changed,
                    "url": case.get("url") or "",
                    "source": source,
                    "will_update": new_status == "PASS" and changed,
                }
                if not case.get("run_id"):
                    no_run_id.append(item)
                    continue
                if new_status != "PASS":
                    skipped.append({**item, "skip_reason": "non_pass"})
                    continue
                if current == "PASS":
                    skipped.append({**item, "skip_reason": "already_present"})
                    continue
                matches.append(item)

        triage = self._build_failure_triage(
            execution_key=execution_key,
            statuses=triage_statuses,
            by_map=by_map,
            report_meta=triage_meta,
            execution_cases=cases,
        )

        return {
            "ok": True,
            "execution": execution_key,
            "mode": "excel_pass_only",
            "files": [report.to_dict()],
            "parsed_map_ids": len(excel_statuses),
            "matches": matches,
            "match_count": len(matches),
            "skipped": skipped,
            "skipped_count": len(skipped),
            "unmatched_html": unmatched_html,
            "unmatched_count": len(unmatched_html),
            "no_run_id": no_run_id,
            "no_run_id_count": len(no_run_id),
            "parse_errors": [
                {"source": report.source, "error": err} for err in report.errors
            ],
            "status_counts": {
                status: sum(1 for s in excel_statuses.values() if s == status)
                for status in sorted(set(excel_statuses.values()))
            },
            **triage,
        }

    def preview_zip_pass_import(
        self,
        execution_key: str,
        *,
        zip_bytes: bytes | None = None,
        zip_name: str = "upload.zip",
        zip_files: list[tuple[str, bytes]] | None = None,
    ) -> dict[str, Any]:
        """Parse one or more ZIPs of HTML reports; later timestamp wins.

        Multiple ZIPs are processed in one preview. Duplicate ZIP/HTML files
        (same basename or content hash) are dropped. Never proposes an update
        when Xray already has PASS.
        """
        if not execution_key:
            raise JiraError("execution is required")
        files = list(zip_files or [])
        if zip_bytes:
            files.append((zip_name or "upload.zip", zip_bytes))
        if not files:
            raise JiraError("ZIP file is required")

        reports, skipped_dupes = parse_zip_files(files)
        # PASS apply: later timestamp wins. Triage: worst status (FAIL) wins.
        html_statuses, chrono_meta = merge_reports_chronological_detailed(reports)
        triage_statuses, triage_meta = merge_reports_detailed(reports)
        winning_sources = {
            mid: (info.get("source") or "") for mid, info in chrono_meta.items()
        }
        cases = self.load_execution_cases(
            execution_key, technology="", force_refresh=True
        )

        by_map: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for case in cases:
            map_id = normalize_map_id(case.get("test_src_map_id") or "")
            if map_id:
                by_map[map_id].append(case)

        matches: list[dict[str, Any]] = []
        skipped: list[dict[str, Any]] = []
        unmatched_html: list[dict[str, Any]] = []
        no_run_id: list[dict[str, Any]] = []

        for map_id, new_status in sorted(html_statuses.items()):
            case_list = by_map.get(map_id) or []
            source = winning_sources.get(map_id) or ""
            if not case_list:
                unmatched_html.append(
                    {
                        "test_src_map_id": map_id,
                        "html_status": new_status,
                        "source": source,
                    }
                )
                continue
            for case in case_list:
                current = self._normalize_status(case.get("status") or "TODO")
                changed = current != new_status
                item = {
                    "test_src_map_id": map_id,
                    "key": case.get("key"),
                    "summary": case.get("summary") or "",
                    "run_id": case.get("run_id"),
                    "current_status": current,
                    "html_status": new_status,
                    "changed": changed,
                    "url": case.get("url") or "",
                    "source": source,
                    "will_update": new_status == "PASS" and changed,
                }
                if not case.get("run_id"):
                    no_run_id.append(item)
                    continue
                if new_status != "PASS":
                    skipped.append({**item, "skip_reason": "non_pass"})
                    continue
                # Already has this result in Xray/TestDeck — leave untouched.
                if current == "PASS":
                    skipped.append({**item, "skip_reason": "already_present"})
                    continue
                matches.append(item)

        parse_errors = []
        for report in reports:
            parse_errors.extend(
                {"source": report.source, "error": err} for err in report.errors
            )

        ordered_files = sorted(
            reports,
            key=lambda r: (r.timestamp if r.timestamp is not None else -1, r.source or ""),
        )

        triage = self._build_failure_triage(
            execution_key=execution_key,
            statuses=triage_statuses,
            by_map=by_map,
            report_meta=triage_meta,
            execution_cases=cases,
        )

        return {
            "ok": True,
            "execution": execution_key,
            "mode": "zip_pass_only",
            "zip_count": len(files),
            "files": [r.to_dict() for r in ordered_files],
            "parsed_map_ids": len(html_statuses),
            "matches": matches,
            "match_count": len(matches),
            "skipped": skipped,
            "skipped_count": len(skipped),
            "unmatched_html": unmatched_html,
            "unmatched_count": len(unmatched_html),
            "no_run_id": no_run_id,
            "no_run_id_count": len(no_run_id),
            "parse_errors": parse_errors,
            "duplicate_files_skipped": skipped_dupes,
            "duplicate_files_skipped_count": len(skipped_dupes),
            "status_counts": {
                status: sum(1 for s in html_statuses.values() if s == status)
                for status in sorted(set(html_statuses.values()))
            },
            **triage,
        }

    def _apply_trcf_entry(self, catalog: dict[str, dict[str, Any]], raw: dict[str, Any]) -> str:
        """Merge a discovered TRCF onto the known Test Details catalog only."""
        name = str(
            raw.get("name") or raw.get("label") or raw.get("customFieldName") or ""
        ).strip()
        field_id = raw.get("customFieldId") or raw.get("cfId") or raw.get("id")
        if field_id is not None and not is_valid_trcf_id(field_id):
            return ""
        # Only keep fields we intentionally show (ignore sdk_build_num, job_type, …).
        key = match_field_key(name) if name else ""
        if not key or key not in catalog:
            return ""
        if is_valid_trcf_id(field_id):
            catalog[key]["id"] = int(field_id)
        options = raw.get("options") or raw.get("allowedValues") or raw.get("values") or []
        for opt in options:
            if isinstance(opt, dict):
                text = normalize_trcf_value(
                    opt.get("value") or opt.get("name") or opt.get("label") or opt
                )
            else:
                text = normalize_trcf_value(opt)
            if text and text not in catalog[key]["options"]:
                catalog[key]["options"].append(text)
        text = normalize_trcf_value(raw.get("value"))
        if text and text.lower() != "none":
            if not catalog[key]["current"]:
                catalog[key]["current"] = text
            if text not in catalog[key]["options"]:
                catalog[key]["options"].append(text)
        return key

    def get_test_run_custom_field_catalog(
        self,
        execution_key: str = "",
        *,
        force_refresh: bool = False,
        cache_only: bool = False,
    ) -> dict[str, Any]:
        """Discover Test Run custom fields from Xray (generic) for Results Update."""
        catalog = {item["key"]: item for item in empty_field_catalog()}
        discovered_from: list[str] = []
        remote_names: list[str] = []
        project_keys: list[str] = []
        if execution_key and "-" in execution_key:
            project_keys.append(execution_key.rsplit("-", 1)[0])
        project_keys.extend(
            [
                settings.JIRA_PROJECT_KEY,
                getattr(settings, "JIRA_TEST_PROJECT_KEY", "") or "",
            ]
        )

        cache_key = f"trcf_catalog_v5:{(execution_key or 'none')}:{','.join(project_keys)}"
        if not force_refresh:
            cached = cache.get(cache_key)
            if isinstance(cached, dict):
                return cached
            if cache_only:
                return {
                    "ok": True,
                    "execution": execution_key or "",
                    "fields": list(catalog.values()),
                    "mapped_ids": 0,
                    "ready": False,
                    "discovered_from": [],
                    "cache_only": True,
                }
        else:
            cache.delete(cache_key)

        # 0) Optional .env overrides (fallback only; discovery should win).
        for key, raw_id in (getattr(settings, "XRAY_TRCF_IDS", {}) or {}).items():
            if key in catalog and str(raw_id or "").strip():
                catalog[key]["id"] = str(raw_id).strip()
        if any(catalog[k]["id"] for k in catalog):
            discovered_from.append("env")

        probe_log: list[dict[str, Any]] = []
        settings_fields: list[dict[str, Any]] = []

        # 1) Primary on Silabs: GET testRunValues?testRunId= for runs in the execution.
        #    This is the same resource used to write values and usually includes ids (+ names).
        if execution_key:
            try:
                samples = self.xray.get_test_execution_tests(
                    execution_key, detailed=True, limit=3, page=1
                )
            except JiraError:
                samples = []
            if samples:
                discovered_from.append("execution_sample")
            for sample in samples[:3]:
                run_id = sample.get("id") or sample.get("testRunId")
                blobs: list[Any] = [sample]
                if run_id:
                    try:
                        values = self.xray.get_test_run_custom_field_values(run_id)
                        if values:
                            discovered_from.append("testRunValues")
                            blobs.append(values)
                            probe_log.append(
                                {
                                    "path": f"testRunValues?testRunId={run_id}",
                                    "status": 200,
                                    "ok": True,
                                    "count": len(values),
                                }
                            )
                    except JiraError as exc:
                        probe_log.append(
                            {
                                "path": f"testRunValues?testRunId={run_id}",
                                "status": exc.status_code,
                                "ok": False,
                                "count": 0,
                            }
                        )
                    try:
                        blobs.append(self.xray.get_test_run(run_id))
                    except JiraError:
                        pass
                for blob in blobs:
                    for raw in extract_named_ids(blob):
                        if raw.get("name"):
                            remote_names.append(str(raw["name"]))
                        self._apply_trcf_entry(catalog, raw)
                    if isinstance(blob, list):
                        for raw in blob:
                            if isinstance(raw, dict):
                                if raw.get("name"):
                                    remote_names.append(str(raw["name"]))
                                self._apply_trcf_entry(catalog, raw)
                if sum(1 for item in catalog.values() if item.get("id") is not None) >= 8:
                    break

        # 2) Project settings (admin endpoints; often 404 for normal users).
        mapped_so_far = sum(1 for item in catalog.values() if item.get("id") is not None)
        if mapped_so_far < 3:
            try:
                settings_fields, settings_probe = self.xray.list_test_run_custom_field_settings(
                    project_keys
                )
                probe_log.extend(settings_probe or [])
            except JiraError:
                settings_fields = []
            if settings_fields:
                discovered_from.append("settings")
                for raw in settings_fields:
                    if not isinstance(raw, dict):
                        continue
                    name = raw.get("name") or raw.get("label") or ""
                    if name:
                        remote_names.append(str(name))
                    self._apply_trcf_entry(catalog, raw)
                for raw in extract_named_ids(settings_fields):
                    if raw.get("name"):
                        remote_names.append(str(raw["name"]))
                    self._apply_trcf_entry(catalog, raw)

        # Keep seed options first, then discovered uniqueness (skip junk values).
        for key, item in catalog.items():
            seeds = SEED_OPTIONS.get(key) or []
            merged: list[str] = []
            for value in seeds + item["options"]:
                text = normalize_trcf_value(value)
                if text and text not in merged:
                    merged.append(text)
            item["options"] = merged
            item["current"] = normalize_trcf_value(item.get("current"))

        # UI: only the curated Test Details fields (never sdk_build_num / job_type / …).
        fields: list[dict[str, Any]] = []
        for item in empty_field_catalog():
            live = catalog.get(item["key"]) or item
            field_id = live.get("id")
            if field_id is not None and not is_valid_trcf_id(field_id):
                field_id = None
            fields.append(
                {
                    "key": item["key"],
                    "label": item["label"],
                    "id": int(field_id) if is_valid_trcf_id(field_id) else None,
                    "options": live.get("options") or item.get("options") or [],
                    "current": live.get("current") or "",
                }
            )

        seen_names: list[str] = []
        for name in remote_names:
            if name not in seen_names:
                seen_names.append(name)

        result = {
            "ok": True,
            "execution": execution_key or "",
            "project_keys": [k for k in project_keys if k],
            "fields": fields,
            "discovered_from": list(dict.fromkeys(discovered_from)),
            "mapped_ids": sum(1 for f in fields if f.get("id") is not None),
            "raw_settings_count": len(settings_fields),
            "remote_field_names": seen_names[:40],
            "probe_log": probe_log[:30],
            "ready": sum(1 for f in fields if f.get("id") is not None) > 0,
        }
        ttl = max(settings.JIRA_CACHE_SECONDS, 600) if result["mapped_ids"] else 60
        cache.set(cache_key, result, ttl)
        remember_key(cache_key)
        return result

    def build_custom_field_payload(
        self, selected: dict[str, str] | None, catalog: list[dict[str, Any]] | None = None
    ) -> tuple[list[dict[str, Any]], list[str]]:
        """Map UI selections to Xray [{id,value}].

        Returns (payload, skipped_labels). Fields without mapped IDs are skipped
        so known IDs (Host Platform, etc.) still post.
        """
        selected = selected or {}
        if catalog is None:
            catalog = empty_field_catalog()
        by_key = {item["key"]: item for item in catalog}
        by_id = {
            str(item.get("id")): item
            for item in catalog
            if item.get("id") is not None
        }
        payload: list[dict[str, Any]] = []
        skipped: list[str] = []
        for key, value in selected.items():
            text = str(value or "").strip()
            if not text or text.lower() in {"none", "—", "-"}:
                continue
            meta = by_key.get(key) or by_id.get(str(key)) or {}
            field_id = meta.get("id")
            label = meta.get("label") or key
            if field_id is None or not is_valid_trcf_id(field_id):
                skipped.append(str(label))
                continue
            payload.append({"id": int(field_id), "value": text})
        return payload, skipped

    def apply_html_import(
        self,
        execution_key: str,
        updates: list[dict[str, Any]],
        *,
        pass_only: bool = False,
        custom_fields: list[dict[str, Any]] | None = None,
        progress_cb: Any | None = None,
    ) -> dict[str, Any]:
        """Apply per-run status updates produced by preview_html_import."""
        if not execution_key:
            raise JiraError("execution is required")
        if not updates:
            raise JiraError("No updates to apply")

        def _progress(payload: dict[str, Any]) -> None:
            if callable(progress_cb):
                try:
                    progress_cb(payload)
                except Exception:
                    pass

        custom_fields = XrayClient._normalize_custom_fields(custom_fields or [])
        already_pass: set[str] = set()

        if pass_only:
            updates = [
                item
                for item in updates
                if self._normalize_status(item.get("status") or "") == "PASS"
            ]
            if not updates:
                raise JiraError("No PASS updates to apply (failed/skipped cases are ignored)")

            # Re-check live execution: never overwrite PASS status.
            # Still allow custom-field writes on already-PASS runs.
            _progress(
                {
                    "phase": "preparing",
                    "message": "Checking current Xray statuses…",
                }
            )
            live_cases = self.load_execution_cases(
                execution_key, technology="", force_refresh=False
            )
            already_pass = {
                str(c.get("run_id"))
                for c in live_cases
                if c.get("run_id")
                and self._normalize_status(c.get("status") or "") == "PASS"
            }
            if not custom_fields:
                updates = [
                    item
                    for item in updates
                    if str(item.get("run_id") or "") not in already_pass
                ]
                if not updates:
                    raise JiraError(
                        "Nothing to post — selected cases already have PASS in Xray"
                    )
            elif not updates:
                raise JiraError("No updates to apply")

        from concurrent.futures import ThreadPoolExecutor, as_completed

        updated: list[dict[str, Any]] = []
        failed: list[dict[str, Any]] = []
        total = len(updates)
        done = 0
        _progress(
            {
                "phase": "updating",
                "message": f"Updating 0/{total}…",
                "total": total,
                "done": 0,
                "left": total,
                "updated_count": 0,
                "failed_count": 0,
            }
        )

        # Silabs Xray: set each TRCF via
        # /customFields/testRunValues?testRunId=&customFieldId= with JSON string body.
        def _one(item: dict[str, Any]) -> tuple[dict[str, Any], Exception | None]:
            run_id = item.get("run_id")
            status = self._normalize_status(item.get("status") or "")
            try:
                if not run_id:
                    raise JiraError("run_id is required")
                if pass_only and status != "PASS":
                    raise JiraError("Only PASS updates are allowed for ZIP import")
                if status not in EDITABLE_STATUSES and status not in STATUS_ALIASES.values():
                    raise JiraError(f"Unsupported status: {status}")
                # Already PASS → write custom fields only (do not re-PUT status).
                status_to_set = None if str(run_id) in already_pass else status
                if status_to_set is None and not custom_fields:
                    raise JiraError("Already PASS — nothing to update")
                self.xray.update_test_run(
                    run_id,
                    status=status_to_set,
                    custom_fields=custom_fields or None,
                )
                return (
                    {
                        "run_id": str(run_id),
                        "status": status,
                        "key": item.get("key") or "",
                        "test_src_map_id": item.get("test_src_map_id") or "",
                        "custom_fields_count": len(custom_fields),
                        "status_skipped": status_to_set is None,
                    },
                    None,
                )
            except Exception as exc:  # noqa: BLE001
                return (
                    {
                        "run_id": str(run_id or ""),
                        "status": status,
                        "key": item.get("key") or "",
                        "test_src_map_id": item.get("test_src_map_id") or "",
                    },
                    exc,
                )

        with ThreadPoolExecutor(max_workers=6) as pool:
            futures = [pool.submit(_one, item) for item in updates]
            for fut in as_completed(futures):
                payload, err = fut.result()
                if err is None:
                    updated.append(payload)
                else:
                    failed.append({**payload, "error": str(err)})
                done += 1
                left = max(0, total - done)
                label = (
                    payload.get("key")
                    or payload.get("test_src_map_id")
                    or payload.get("run_id")
                    or ""
                )
                _progress(
                    {
                        "phase": "updating",
                        "message": f"Updating {done}/{total}…",
                        "current": str(label),
                        "total": total,
                        "done": done,
                        "left": left,
                        "updated_count": len(updated),
                        "failed_count": len(failed),
                    }
                )

        self._bust_execution_case_cache(execution_key)

        return {
            "ok": not failed,
            "execution": execution_key,
            "updated": updated,
            "failed": failed,
            "updated_count": len(updated),
            "failed_count": len(failed),
            "custom_fields_count": len(custom_fields),
        }

    def get_execution_dashboard(
        self,
        execution_key: str,
        section_path: str = "",
        status_filter: str = "",
        search: str = "",
        technology: str | None = None,
        page: int = 1,
        force_refresh: bool = False,
    ) -> dict[str, Any]:
        field_map = self.resolve_fields()
        technology = technology if technology is not None else settings.DEFAULT_TECHNOLOGY

        exec_fields_key = f"exec_issue:{execution_key}"
        execution = cache.get(exec_fields_key)
        if execution is None or force_refresh:
            exec_field_list = [
                "summary",
                "status",
                "priority",
                "assignee",
                "reporter",
                "created",
                "updated",
                "labels",
                "fixVersions",
            ]
            for fid in ("test_plan", "test_environments", "stack_name", "feature_name", "tech_area"):
                mapped = field_map.get(fid)
                if mapped:
                    exec_field_list.append(mapped)
            execution = self.jira.get_issue(execution_key, fields=",".join(exec_field_list))
            cache.set(exec_fields_key, execution, settings.JIRA_CACHE_SECONDS)
            remember_key(exec_fields_key)

        exec_fields = execution.get("fields") or {}
        cases = self.load_execution_cases(
            execution_key, technology=technology, force_refresh=force_refresh
        )

        # Sidebar tree respects status/search so empty sections are hidden.
        tree_cases = self._filter_cases(cases, "", status_filter, search)
        section_tree = self._build_section_tree(tree_cases)
        filtered = self._filter_cases(cases, section_path, status_filter, search)
        summary = self._summarize_statuses(filtered)
        overall = self._summarize_statuses(cases)
        paging = self.paginate_cases(filtered, page=page)
        coverage = self._build_coverage(filtered)

        return {
            "execution": {
                "key": execution.get("key"),
                "summary": exec_fields.get("summary") or "",
                "status": ((exec_fields.get("status") or {}).get("name") or ""),
                "priority": ((exec_fields.get("priority") or {}).get("name") or ""),
                "assignee": self._user_name(exec_fields.get("assignee")),
                "reporter": self._user_name(exec_fields.get("reporter")),
                "created": exec_fields.get("created"),
                "updated": exec_fields.get("updated"),
                "labels": exec_fields.get("labels") or [],
                "fixversions": [v.get("name") for v in (exec_fields.get("fixVersions") or [])],
                "url": self.jira.browse_url(execution_key),
                "test_plan": self._extract_field(exec_fields, field_map.get("test_plan")),
                "environments": self._extract_field(
                    exec_fields, field_map.get("test_environments")
                ),
                "stack_name": self._extract_field(exec_fields, field_map.get("stack_name")),
                "feature_name": self._extract_field(exec_fields, field_map.get("feature_name")),
                "tech_area": self._extract_field(exec_fields, field_map.get("tech_area")),
            },
            "section_path": section_path,
            "section_name": section_path.split("/")[-1] if section_path else "All Tests",
            "sections": [n.to_dict() for n in section_tree],
            "summary": summary.to_dict(),
            "overall_summary": overall.to_dict(),
            "cases": paging["page_cases"],
            "case_count": len(filtered),
            "total_cases": len(tree_cases),
            "unfiltered_total": len(cases),
            "pagination": paging,
            "defects": [],
            "coverage": coverage,
            "editable_statuses": EDITABLE_STATUSES,
            "filters": {
                "status": status_filter,
                "search": search,
                "technology": technology or "",
            },
            "status_colors": STATUS_COLORS,
            "technology": technology or "",
        }

    def get_test_plan_dashboard(
        self,
        plan_key: str,
        section_path: str = "",
        status_filter: str = "",
        search: str = "",
        technology: str | None = None,
        execution_key: str = "",
        page: int = 1,
        force_refresh: bool = False,
        include_execution_list: bool = True,
    ) -> dict[str, Any]:
        """
        TestRail-style plan view:
        - section tree
        - aggregate + per-case status from selected execution (Option B)
        """
        field_map = self.resolve_fields()
        technology = technology if technology is not None else settings.DEFAULT_TECHNOLOGY

        plan_cache_key = f"plan_issue:{plan_key}"
        plan_issue = None if force_refresh else cache.get(plan_cache_key)
        if plan_issue is None:
            plan_field_list = [
                "summary",
                "status",
                "priority",
                "assignee",
                "reporter",
                "created",
                "updated",
                "labels",
                "fixVersions",
            ]
            for key in ("stack_name", "release_name"):
                fid = field_map.get(key) or settings.XRAY_FIELD_MAP.get(key)
                if fid and fid not in plan_field_list:
                    plan_field_list.append(fid)
            plan_issue = self.jira.get_issue(plan_key, fields=",".join(plan_field_list))
            cache.set(plan_cache_key, plan_issue, settings.JIRA_CACHE_SECONDS)
            remember_key(plan_cache_key)
        plan_fields = plan_issue.get("fields") or {}

        execution_rows: list[dict[str, Any]] = []
        if include_execution_list:
            exec_cache_key = f"plan_execs:{plan_key}"
            executions = None if force_refresh else cache.get(exec_cache_key)
            if executions is None:
                executions = self.xray.get_test_plan_executions(plan_key)
                cache.set(exec_cache_key, executions, settings.JIRA_CACHE_SECONDS)
                remember_key(exec_cache_key)

            execution_rows = []
            for item in executions or []:
                if not isinstance(item, dict) or not item.get("key"):
                    continue
                envs = item.get("testEnvironments") or []
                if isinstance(envs, str):
                    env_text = envs
                elif isinstance(envs, list):
                    parts = []
                    for env in envs:
                        if isinstance(env, str):
                            parts.append(env)
                        elif isinstance(env, dict):
                            parts.append(
                                str(
                                    env.get("name")
                                    or env.get("value")
                                    or env.get("key")
                                    or ""
                                )
                            )
                        elif env is not None:
                            parts.append(str(env))
                    env_text = ", ".join([p for p in parts if p])
                else:
                    env_text = str(envs)
                execution_rows.append(
                    {
                        "key": item.get("key"),
                        "summary": item.get("summary") or "",
                        "environments": env_text,
                        "url": self.jira.browse_url(item["key"]),
                    }
                )

            def _exec_num(row: dict[str, Any]) -> int:
                try:
                    return int(str(row["key"]).split("-")[-1])
                except Exception:
                    return 0

            execution_rows.sort(key=_exec_num, reverse=True)
        elif execution_key:
            execution_rows = [
                {
                    "key": execution_key,
                    "summary": "",
                    "environments": "",
                    "url": self.jira.browse_url(execution_key),
                }
            ]

        # Fast path: selected execution uses cached parallel loader
        if execution_key:
            source = "execution"
            cases = self.load_execution_cases(
                execution_key, technology=technology, force_refresh=force_refresh
            )
        else:
            source = "plan"
            safe_tech = (technology or "").replace('"', '\\"')
            jql = f"issue in testPlanTests({plan_key})"
            if technology:
                jql += f' AND Technology = "{safe_tech}"'
            jql += " ORDER BY key ASC"
            fields = ["summary", "priority", "assignee", "labels", "components"]
            for fid in field_map.values():
                if fid and fid not in fields:
                    fields.append(fid)
            issues = self.jira.search_all(
                jql=jql,
                fields=fields,
                page_size=200,
                hard_limit=settings.XRAY_HARD_LIMIT,
            )
            cases = []
            for issue in issues:
                f = issue.get("fields") or {}
                key = issue.get("key")
                path = self._extract_field(f, field_map.get("test_repo_path")) or self._path_from_labels(
                    f.get("labels") or []
                )
                cases.append(
                    {
                        "key": key,
                        "run_id": None,
                        "summary": f.get("summary") or "",
                        "status": "TODO",
                        "assignee": self._user_name(f.get("assignee")),
                        "priority": ((f.get("priority") or {}).get("name") or ""),
                        "labels": f.get("labels") or [],
                        "section_path": self._clean_path(path) or "Uncategorized",
                        "technology": self._extract_field(f, field_map.get("technology"))
                        or technology,
                        "test_src_map_id": self._extract_field(
                            f, field_map.get("test_src_map_id")
                        ),
                        "defects": [],
                        "url": self.jira.browse_url(key),
                    }
                )

        # Sidebar tree respects status/search so empty sections are hidden.
        tree_cases = self._filter_cases(cases, "", status_filter, search)
        section_tree = self._build_section_tree(tree_cases)
        filtered = self._filter_cases(cases, section_path, status_filter, search)
        summary = self._summarize_statuses(filtered)
        overall = self._summarize_statuses(cases)
        paging = self.paginate_cases(filtered, page=page)
        coverage = self._build_coverage(filtered)

        return {
            "plan": {
                "key": plan_issue.get("key"),
                "summary": plan_fields.get("summary") or "",
                "status": ((plan_fields.get("status") or {}).get("name") or ""),
                "assignee": self._user_name(plan_fields.get("assignee")),
                "reporter": self._user_name(plan_fields.get("reporter")),
                "created": plan_fields.get("created"),
                "updated": plan_fields.get("updated"),
                "stack_name": self._extract_field(
                    plan_fields,
                    field_map.get("stack_name")
                    or settings.XRAY_FIELD_MAP.get("stack_name"),
                ),
                "release_name": self._extract_field(
                    plan_fields,
                    field_map.get("release_name")
                    or settings.XRAY_FIELD_MAP.get("release_name"),
                ),
                "url": self.jira.browse_url(plan_key),
                "fixversions": [
                    (v.get("name") if isinstance(v, dict) else str(v))
                    for v in (plan_fields.get("fixVersions") or [])
                    if v
                ],
            },
            "executions": execution_rows,
            "selected_execution": execution_key,
            "status_source": source,
            "section_path": section_path,
            "section_name": section_path.split("/")[-1] if section_path else "All Tests",
            "sections": [n.to_dict() for n in section_tree],
            "summary": summary.to_dict(),
            "overall_summary": overall.to_dict(),
            "cases": paging["page_cases"],
            "case_count": len(filtered),
            "total_cases": len(tree_cases),
            "unfiltered_total": len(cases),
            "pagination": paging,
            "defects": [],
            "coverage": coverage,
            "editable_statuses": EDITABLE_STATUSES,
            "filters": {
                "status": status_filter,
                "search": search,
                "technology": technology or "",
            },
            "technology": technology or "",
            "status_colors": STATUS_COLORS,
        }

    def get_overview(self) -> dict[str, Any]:
        executions = self.list_test_executions(limit=12)
        plans = self.list_test_plans(limit=8)
        connection = self.connection_status()
        return {
            "connection": connection,
            "executions": executions,
            "plans": plans,
            "project_key": settings.JIRA_PROJECT_KEY,
            "test_project_key": settings.JIRA_TEST_PROJECT_KEY,
            "jira_base_url": settings.JIRA_BASE_URL,
        }

    def get_repository_tests(
        self,
        folder_hint: str = "",
        search: str = "",
        limit: int = 100,
        technology: str | None = None,
    ) -> dict[str, Any]:
        project = settings.JIRA_TEST_PROJECT_KEY
        issue_type = settings.JIRA_ISSUE_TYPE_TEST
        technology = technology if technology is not None else settings.DEFAULT_TECHNOLOGY
        clauses = [f"project = {project}", f'issuetype = "{issue_type}"']
        selected = self._clean_path(folder_hint)
        # Paths are filtered in-memory; only use non-path hints in JQL.
        if selected and "/" not in selected:
            safe = selected.replace('"', '\\"')
            clauses.append(
                f'(labels = "{safe}" OR summary ~ "{safe}" OR description ~ "{safe}")'
            )
        if technology.strip():
            safe = technology.replace('"', '\\"')
            clauses.append(f'Technology = "{safe}"')
        if search.strip():
            safe = search.replace('"', '\\"')
            clauses.append(f'(key = "{safe}" OR summary ~ "{safe}")')
        jql = " AND ".join(clauses) + " ORDER BY updated DESC"

        field_map = self.resolve_fields()
        fields = [
            "summary",
            "status",
            "priority",
            "assignee",
            "labels",
            "components",
            "updated",
        ]
        for fid in field_map.values():
            if fid and fid not in fields:
                fields.append(fid)

        issues = self.jira.search(jql=jql, fields=fields, max_results=limit).get("issues", [])
        tests = []
        for issue in issues:
            f = issue.get("fields") or {}
            tests.append(
                {
                    "key": issue.get("key"),
                    "summary": f.get("summary") or "",
                    "status": ((f.get("status") or {}).get("name") or ""),
                    "priority": ((f.get("priority") or {}).get("name") or ""),
                    "assignee": self._user_name(f.get("assignee")),
                    "labels": f.get("labels") or [],
                    "components": [c.get("name") for c in (f.get("components") or [])],
                    "section_path": self._extract_field(f, field_map.get("test_repo_path"))
                    or self._path_from_labels(f.get("labels") or []),
                    "feature_name": self._extract_field(f, field_map.get("feature_name")),
                    "stack_name": self._extract_field(f, field_map.get("stack_name")),
                    "test_src_map_id": self._extract_field(
                        f, field_map.get("test_src_map_id")
                    ),
                    "technology": self._extract_field(f, field_map.get("technology")),
                    "url": self.jira.browse_url(issue["key"]),
                }
            )

        sections = self._build_section_tree(tests)
        visible = tests
        if selected:
            visible = [
                t
                for t in tests
                if self._clean_path(t.get("section_path") or "") == selected
                or self._clean_path(t.get("section_path") or "").startswith(selected + "/")
            ]
        return {
            "tests": visible,
            "sections": [n.to_dict() for n in sections],
            "count": len(visible),
            "jql": jql,
            "section_path": selected,
            "technology": technology,
            "filters": {"search": search, "technology": technology},
        }

    # --- helpers ---------------------------------------------------------

    def _normalize_execution_tests(
        self, raw_tests: list[dict[str, Any]], field_map: dict[str, str]
    ) -> list[dict[str, Any]]:
        cases: list[dict[str, Any]] = []
        for idx, item in enumerate(raw_tests, start=1):
            key = (
                item.get("key")
                or item.get("testKey")
                or item.get("issueKey")
                or (item.get("test") or {}).get("key")
            )
            if not key:
                continue

            status_raw = (
                item.get("status")
                or item.get("statusName")
                or (item.get("status") or {})
            )
            if isinstance(status_raw, dict):
                status_raw = status_raw.get("name") or status_raw.get("description") or "TODO"
            status = self._normalize_status(str(status_raw))

            summary = (
                item.get("summary")
                or item.get("testSummary")
                or (item.get("test") or {}).get("summary")
                or ""
            )
            assignee = self._run_assignee(item)

            defects = item.get("defects") or item.get("defectKeys") or []
            if isinstance(defects, str):
                defects = [defects]

            section_path = (
                item.get("testRepositoryPath")
                or item.get("repositoryPath")
                or item.get("folder")
                or ""
            )
            if not section_path and isinstance(item.get("test"), dict):
                section_path = item["test"].get("testRepositoryPath") or ""

            cases.append(
                {
                    "rank": item.get("rank") or idx,
                    "key": key,
                    "summary": summary,
                    "status": status,
                    "status_raw": str(status_raw),
                    "assignee": assignee or "",
                    "test_type": item.get("testType")
                    or item.get("type")
                    or ((item.get("test") or {}).get("testType"))
                    or "",
                    "defects": defects if isinstance(defects, list) else [],
                    "section_path": self._clean_path(section_path),
                    "labels": item.get("labels") or [],
                    "priority": item.get("priority") or "",
                    "url": self.jira.browse_url(key),
                    "iterations": item.get("iterations") or item.get("datasets") or "",
                }
            )
        return cases

    def _bulk_fetch_tests(
        self, keys: list[str], field_map: dict[str, str]
    ) -> dict[str, dict[str, Any]]:
        if not keys:
            return {}

        fields = [
            "summary",
            "priority",
            "assignee",
            "labels",
            "components",
            "issuelinks",
            "status",
        ]
        for fid in field_map.values():
            if fid and fid not in fields:
                fields.append(fid)

        enriched: dict[str, dict[str, Any]] = {}
        # Jira JQL IN clause batches
        chunk_size = 50
        for i in range(0, len(keys), chunk_size):
            chunk = keys[i : i + chunk_size]
            quoted = ", ".join(f'"{k}"' for k in chunk)
            jql = f"key in ({quoted})"
            issues = self.jira.search(jql=jql, fields=fields, max_results=chunk_size).get(
                "issues", []
            )
            for issue in issues:
                f = issue.get("fields") or {}
                path = self._extract_field(f, field_map.get("test_repo_path"))
                if not path:
                    path = self._path_from_labels(f.get("labels") or [])
                defects = self._defect_keys_from_links(f.get("issuelinks") or [])
                enriched[issue["key"]] = {
                    "summary": f.get("summary") or "",
                    "priority": ((f.get("priority") or {}).get("name") or ""),
                    "assignee": self._user_name(f.get("assignee")),
                    "labels": f.get("labels") or [],
                    "components": [c.get("name") for c in (f.get("components") or [])],
                    "section_path": self._clean_path(path),
                    "feature_name": self._extract_field(f, field_map.get("feature_name")),
                    "stack_name": self._extract_field(f, field_map.get("stack_name")),
                    "tech_area": self._extract_field(f, field_map.get("tech_area")),
                    "technology": self._extract_field(f, field_map.get("technology")),
                    "test_src_map_id": self._extract_field(
                        f, field_map.get("test_src_map_id")
                    ),
                    "testrail_section": self._extract_field(
                        f, field_map.get("testrail_section")
                    ),
                    "defects": defects,
                }
        return enriched

    def _build_section_tree(self, cases: list[dict[str, Any]]) -> list[SectionNode]:
        # path -> cases
        by_path: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for case in cases:
            path = self._clean_path(case.get("section_path") or "Uncategorized")
            case["section_path"] = path
            by_path[path].append(case)

        root_children: dict[str, SectionNode] = {}

        for path, path_cases in sorted(by_path.items()):
            parts = [p for p in path.split("/") if p]
            current_map = root_children
            built_path = []
            for part in parts:
                built_path.append(part)
                full = "/".join(built_path)
                if full not in current_map and part not in {n.name for n in current_map.values()}:
                    # store by full path key in a nested structure via nodes
                    pass

        # Simpler approach: build nested dict then convert
        tree: dict[str, Any] = {}
        for path, path_cases in by_path.items():
            parts = [p for p in path.split("/") if p] or ["Uncategorized"]
            node = tree
            for part in parts:
                node = node.setdefault(part, {"__cases__": [], "__children__": {}})
                node = node["__children__"]
            # attach cases at leaf
            leaf = tree
            for part in parts[:-1]:
                leaf = leaf[part]["__children__"]
            leaf[parts[-1]]["__cases__"].extend(path_cases)

        def convert(name: str, payload: dict[str, Any], parent_path: str) -> SectionNode:
            path = f"{parent_path}/{name}" if parent_path else name
            child_nodes = [
                convert(child_name, child_payload, path)
                for child_name, child_payload in sorted(payload.get("__children__", {}).items())
            ]
            # accumulate all descendant cases
            all_cases = list(payload.get("__cases__", []))
            for child in child_nodes:
                # recount via filtering cases under path
                pass
            descendant_cases = [
                c for c in cases if self._clean_path(c.get("section_path") or "").startswith(path)
            ]
            return SectionNode(
                id=path,
                name=name,
                path=path,
                count=len(descendant_cases),
                children=child_nodes,
                status=self._summarize_statuses(descendant_cases),
            )

        return [
            convert(name, payload, "")
            for name, payload in sorted(tree.items())
        ]

    def _filter_cases(
        self,
        cases: list[dict[str, Any]],
        section_path: str,
        status_filter: str,
        search: str,
    ) -> list[dict[str, Any]]:
        section_path = self._clean_path(section_path)
        status_filter = (status_filter or "").strip().upper()
        search = (search or "").strip().lower()

        result = []
        for case in cases:
            path = self._clean_path(case.get("section_path") or "")
            if section_path and path != section_path and not path.startswith(section_path + "/"):
                continue
            if status_filter and status_filter != "ALL" and case.get("status") != status_filter:
                continue
            if search:
                blob = f"{case.get('key','')} {case.get('summary','')}".lower()
                if search not in blob:
                    continue
            result.append(case)
        return result

    @staticmethod
    def _defect_keys(defects: Any) -> list[str]:
        keys: list[str] = []
        if not defects:
            return keys
        if isinstance(defects, str):
            return [defects] if defects.strip() else []
        if not isinstance(defects, list):
            return [str(defects)]
        for item in defects:
            if isinstance(item, str) and item.strip():
                keys.append(item.strip())
            elif isinstance(item, dict):
                key = item.get("key") or item.get("id") or item.get("defectKey")
                if key:
                    keys.append(str(key))
        return keys

    @staticmethod
    def _cell_list(values: list[Any], *, keep_blanks: bool = False) -> str:
        """Join values as a vertical list inside one Excel cell (newline, no separators)."""
        parts: list[str] = []
        for v in values:
            text = "" if v is None else str(v).strip()
            if text or keep_blanks:
                parts.append(text)
        return "\n".join(parts)

    @staticmethod
    def _autosize_sheet(ws, sample_rows: int = 200, max_width: int = 56) -> None:
        from openpyxl.styles import Alignment

        for col in ws.columns:
            letter = col[0].column_letter
            max_len = 0
            for cell in col[:sample_rows]:
                val = "" if cell.value is None else str(cell.value)
                if "\n" in val:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
                    for line in val.splitlines():
                        if len(line) > max_len:
                            max_len = len(line)
                elif len(val) > max_len:
                    max_len = len(val)
            ws.column_dimensions[letter].width = min(max(10, max_len + 2), max_width)

    @staticmethod
    def _step_field_text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        if isinstance(value, dict):
            # v1: {'raw': '...', 'rendered': '...'}  v2 nested under fields.*.value
            if "raw" in value or "rendered" in value:
                return str(value.get("raw") or value.get("rendered") or "").strip()
            inner = value.get("value")
            if isinstance(inner, dict):
                return str(inner.get("raw") or inner.get("rendered") or "").strip()
            if inner is not None:
                return str(inner).strip()
            return str(value.get("text") or "").strip()
        return str(value).strip()

    def _normalize_test_steps(self, steps: list[dict[str, Any]]) -> list[dict[str, str]]:
        normalized: list[dict[str, str]] = []
        for idx, step in enumerate(steps or [], start=1):
            if not isinstance(step, dict):
                continue
            fields = step.get("fields") if isinstance(step.get("fields"), dict) else {}
            action = self._step_field_text(
                step.get("step") or fields.get("Action") or fields.get("action")
            )
            data = self._step_field_text(
                step.get("data") or fields.get("Data") or fields.get("data")
            )
            expected = self._step_field_text(
                step.get("result")
                or fields.get("Expected Result")
                or fields.get("ExpectedResult")
                or fields.get("expectedResult")
            )
            index = step.get("index") or idx
            normalized.append(
                {
                    "index": str(index),
                    "action": action,
                    "data": data,
                    "expected": expected,
                }
            )
        return normalized

    def _format_steps_block(self, steps: list[dict[str, str]]) -> str:
        lines: list[str] = []
        for step in steps:
            idx = step.get("index") or ""
            action = step.get("action") or ""
            data = step.get("data") or ""
            expected = step.get("expected") or ""
            chunk = [f"{idx}. {action}".strip()]
            if data:
                chunk.append(f"   Data: {data}")
            if expected:
                chunk.append(f"   Expected: {expected}")
            lines.append("\n".join(chunk))
        text = "\n".join(lines).strip()
        if len(text) > 32000:
            return text[:32000] + "…"
        return text

    def _load_steps_by_key(self, keys: list[str]) -> dict[str, list[dict[str, str]]]:
        from concurrent.futures import ThreadPoolExecutor, as_completed

        unique = [k for k in dict.fromkeys(keys) if k]
        out: dict[str, list[dict[str, str]]] = {}
        if not unique:
            return out

        def _one(key: str) -> tuple[str, list[dict[str, str]]]:
            try:
                raw = self.xray.get_test_steps(key)
                return key, self._normalize_test_steps(raw)
            except Exception:
                return key, []

        workers = min(8, max(1, len(unique)))
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futures = [pool.submit(_one, key) for key in unique]
            for fut in as_completed(futures):
                key, steps = fut.result()
                out[key] = steps
        return out

    @staticmethod
    def _milestone_from_execution_name(name: str) -> str:
        """Parse Milestone from TE naming: te-<Milestone>-... (split on '-').

        Example: te-wc_4.1.1_cf-917_IoT_FreeRTOS-NCP-sample_apps-gcc_lto-Bbrd4346a-spi
        → Milestone = wc_4.1.1_cf
        """
        text = (name or "").strip()
        if not text:
            return ""
        # Prefer the te-... token if the summary has extra wording.
        import re

        m = re.search(r"\bte-[^\s,;]+", text, flags=re.IGNORECASE)
        token = m.group(0) if m else text
        parts = [p for p in token.split("-") if p != ""]
        if len(parts) >= 2 and parts[0].lower() == "te":
            return parts[1]
        if len(parts) >= 2:
            return parts[1]
        return ""

    @staticmethod
    def _export_case_headers(*, include_steps: bool) -> list[str]:
        headers = [
            "Test Execution",
            "Test Execution name",
            "Milestone",
            "ID",
            "Case ID",
            "test_src_map_id",
            "Title",
            "References",
            "Reference key",
            "Tech Area",
            "Feature",
            "Sub-Feature",
            "Parameters",
        ]
        if include_steps:
            headers.extend(
                [
                    "Manual steps",
                    "Step actions",
                    "Step data",
                    "Expected results",
                ]
            )
        headers.extend(
            [
                "Status",
                "Assigned To",
                "Assignee key",
                "Priority",
                "Section",
                "Technology",
                "Case feature",
                "Stack",
                "TestRail section",
                "Labels",
                "Components",
                "Defect count",
                "Defects",
                "Defect statuses",
                "Defect priorities",
                "Has defects",
                "Run ID",
                "Rank",
                "URL",
            ]
        )
        return headers

    @staticmethod
    def _issue_key_from_reference(value: str) -> str:
        import re

        raw = (value or "").strip()
        if not raw:
            return ""
        # URL .../browse/RSCDEV-23314 or bare key
        m = re.search(r"([A-Za-z][A-Za-z0-9_]+-\d+)", raw)
        return (m.group(1) if m else "").upper()

    @staticmethod
    def _split_reference_fields(text: str) -> dict[str, str]:
        """Split reference description/summary on ':' into Tech Area / Feature / Sub-Feature / Parameters."""
        parts = [p.strip() for p in (text or "").split(":")]
        parts = [p for p in parts if p]  # drop empty segments from "A : B"
        # Re-split keeping empties only if original had structure like "a::b"
        if not parts and (text or "").strip():
            parts = [p.strip() for p in (text or "").split(":")]
        while len(parts) < 4:
            parts.append("")
        if len(parts) > 4:
            parts = [parts[0], parts[1], parts[2], ": ".join(parts[3:])]
        return {
            "tech_area": parts[0],
            "feature": parts[1],
            "sub_feature": parts[2],
            "parameters": parts[3],
        }

    def _fetch_reference_issue_fields(
        self, keys: list[str]
    ) -> dict[str, dict[str, Any]]:
        """Fetch summary/description for References targets (e.g. RSCDEV-*)."""
        out: dict[str, dict[str, Any]] = {}
        unique = [k for k in dict.fromkeys(keys) if k and self._looks_like_issue_key(k)]
        if not unique:
            return out
        fields = ["summary", "description", "status", "priority"]
        batch_size = 50
        for i in range(0, len(unique), batch_size):
            batch = unique[i : i + batch_size]
            quoted = ", ".join(f'"{k}"' for k in batch)
            issues: list[dict[str, Any]] = []
            try:
                issues = self.jira.search_all(
                    jql=f"key in ({quoted})",
                    fields=fields,
                    page_size=min(100, len(batch)),
                    hard_limit=len(batch) + 5,
                )
            except JiraError:
                for key in batch:
                    try:
                        issues.append(
                            self.jira.get_issue(key, fields=",".join(fields))
                        )
                    except JiraError:
                        continue
            for issue in issues or []:
                key = (issue.get("key") or "").strip()
                if not key:
                    continue
                f = issue.get("fields") or {}
                desc = self._plain_description(f.get("description"))
                summary = (f.get("summary") or "").strip()
                # Prefer description when it contains ':' structure; else summary
                # (many RSCDEV epics store the taxonomy only in summary).
                source = desc if desc and ":" in desc else (summary or desc)
                split = self._split_reference_fields(source)
                out[key] = {
                    "key": key,
                    "summary": summary,
                    "description": desc,
                    "status": ((f.get("status") or {}).get("name") or ""),
                    "priority": ((f.get("priority") or {}).get("name") or ""),
                    "url": self.jira.browse_url(key),
                    **split,
                }
        return out

    def _export_case_row(
        self,
        case: dict[str, Any],
        *,
        execution_key: str,
        execution_summary: str,
        steps_by_key: dict[str, list[dict[str, str]]],
        include_steps: bool,
        defect_details: dict[str, dict[str, Any]] | None = None,
        reference_details: dict[str, dict[str, Any]] | None = None,
    ) -> list[Any]:
        defect_keys = self._defect_keys(case.get("defects"))
        details = defect_details or {}
        status_bits: list[str] = []
        priority_bits: list[str] = []
        for key in defect_keys:
            info = details.get(key) or {}
            st = (info.get("status") or "").strip()
            pr = (info.get("priority") or "").strip()
            status_bits.append(st)
            priority_bits.append(pr)
        labels = case.get("labels") or []
        components = case.get("components") or []
        ref_raw = case.get("references") or ""
        ref_key = self._issue_key_from_reference(ref_raw)
        ref_info = (reference_details or {}).get(ref_key) or {}
        row: list[Any] = [
            execution_key or "",
            execution_summary or "",
            self._milestone_from_execution_name(execution_summary),
            case.get("key") or "",
            case.get("case_id") or "",
            case.get("test_src_map_id") or "",
            case.get("summary") or "",
            ref_raw,
            ref_key or ref_info.get("key") or "",
            ref_info.get("tech_area") or "",
            ref_info.get("feature") or "",
            ref_info.get("sub_feature") or "",
            ref_info.get("parameters") or "",
        ]
        if include_steps:
            steps = steps_by_key.get(case.get("key") or "") or []
            actions = "\n".join(
                f"{s.get('index')}. {s.get('action')}" for s in steps if s.get("action")
            )
            data_vals = "\n".join(
                f"{s.get('index')}. {s.get('data')}" for s in steps if s.get("data")
            )
            expected_vals = "\n".join(
                f"{s.get('index')}. {s.get('expected')}"
                for s in steps
                if s.get("expected")
            )
            row.extend(
                [
                    self._format_steps_block(steps),
                    actions[:32000],
                    data_vals[:32000],
                    expected_vals[:32000],
                ]
            )
        row.extend(
            [
                case.get("status") or "",
                case.get("assignee") or "",
                case.get("assignee_key") or "",
                case.get("priority") or "",
                case.get("section_path") or "",
                case.get("technology") or "",
                case.get("feature_name") or "",
                case.get("stack_name") or "",
                case.get("testrail_section") or "",
                ", ".join(str(x) for x in labels if x),
                ", ".join(str(x) for x in components if x),
                len(defect_keys),
                self._cell_list(defect_keys),
                self._cell_list(status_bits, keep_blanks=True),
                self._cell_list(priority_bits, keep_blanks=True),
                "Yes" if defect_keys else "No",
                case.get("run_id") or "",
                case.get("rank") if case.get("rank") is not None else "",
                case.get("url") or "",
            ]
        )
        return row

    def _write_export_workbook(
        self,
        *,
        meta_rows: list[list[Any]],
        case_entries: list[dict[str, Any]],
        include_steps: bool,
        executions_sheet: list[dict[str, Any]] | None = None,
    ) -> bytes:
        from io import BytesIO

        from openpyxl import Workbook

        headers = self._export_case_headers(include_steps=include_steps)
        all_cases = [e["case"] for e in case_entries]
        summary = self._summarize_statuses(all_cases)
        defect_rows = self._collect_defects(all_cases)
        defect_details = {d["key"]: d for d in defect_rows if d.get("key")}
        ref_keys = [
            self._issue_key_from_reference(c.get("references") or "")
            for c in all_cases
        ]
        reference_details = self._fetch_reference_issue_fields(ref_keys)

        steps_by_key: dict[str, list[dict[str, str]]] = {}
        if include_steps:
            steps_by_key = self._load_steps_by_key(
                [c.get("key") or "" for c in all_cases]
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Cases"
        ws.append(headers)
        for entry in case_entries:
            ws.append(
                self._export_case_row(
                    entry["case"],
                    execution_key=entry.get("execution_key") or "",
                    execution_summary=entry.get("execution_summary") or "",
                    steps_by_key=steps_by_key,
                    include_steps=include_steps,
                    defect_details=defect_details,
                    reference_details=reference_details,
                )
            )
        self._autosize_sheet(ws)

        fail_ws = wb.create_sheet("Failed cases")
        fail_ws.append(headers)
        for entry in case_entries:
            case = entry["case"]
            if self._normalize_status(case.get("status") or "") != "FAIL":
                continue
            fail_ws.append(
                self._export_case_row(
                    case,
                    execution_key=entry.get("execution_key") or "",
                    execution_summary=entry.get("execution_summary") or "",
                    steps_by_key=steps_by_key,
                    include_steps=include_steps,
                    defect_details=defect_details,
                    reference_details=reference_details,
                )
            )
        self._autosize_sheet(fail_ws)

        if include_steps:
            steps_ws = wb.create_sheet("Test steps")
            steps_ws.append(
                [
                    "Test Execution",
                    "Test Execution name",
                    "ID",
                    "Case ID",
                    "Title",
                    "Status",
                    "Step #",
                    "Action",
                    "Data",
                    "Expected Result",
                ]
            )
            for entry in case_entries:
                case = entry["case"]
                steps = steps_by_key.get(case.get("key") or "") or []
                base = [
                    entry.get("execution_key") or "",
                    entry.get("execution_summary") or "",
                    case.get("key") or "",
                    case.get("case_id") or "",
                    case.get("summary") or "",
                    case.get("status") or "",
                ]
                if not steps:
                    steps_ws.append(base + ["", "", "", ""])
                    continue
                for step in steps:
                    steps_ws.append(
                        base
                        + [
                            step.get("index") or "",
                            step.get("action") or "",
                            step.get("data") or "",
                            step.get("expected") or "",
                        ]
                    )
            self._autosize_sheet(steps_ws)

        if executions_sheet is not None:
            # Defects rolled up per Test Execution (from cases in that run).
            defects_by_exec: dict[str, list[str]] = defaultdict(list)
            for entry in case_entries:
                exec_key = (entry.get("execution_key") or "").strip()
                if not exec_key:
                    continue
                for dkey in self._defect_keys(entry.get("case", {}).get("defects")):
                    if dkey not in defects_by_exec[exec_key]:
                        defects_by_exec[exec_key].append(dkey)
            # Fallback when case_entries lack execution_key but sheet items carry cases.
            for item in executions_sheet:
                exec_key = (item.get("key") or "").strip()
                if not exec_key or defects_by_exec.get(exec_key):
                    continue
                for case in item.get("cases") or []:
                    for dkey in self._defect_keys(case.get("defects")):
                        if dkey not in defects_by_exec[exec_key]:
                            defects_by_exec[exec_key].append(dkey)

            exec_ws = wb.create_sheet("Test Executions")
            # One row per execution; Jiras stacked with newlines in a single cell.
            exec_ws.append(
                [
                    "Test Execution",
                    "Test Execution name",
                    "Milestone",
                    "Total",
                    "Coverage",
                    "Coverage %",
                    "Passed",
                    "Passed %",
                    "Failed",
                    "Failed %",
                    "Retest",
                    "Retest %",
                    "Untested",
                    "Untested %",
                    "Blocked",
                    "Blocked %",
                    "Defect count",
                    "Jiras",
                    "URL",
                ]
            )
            for item in executions_sheet:
                metrics = item.get("metrics")
                if not isinstance(metrics, dict):
                    metrics = self._metrics_from_counts(
                        total=int(item.get("total") or 0),
                        passed=int(item.get("passed") or 0),
                        failed=int(item.get("failed") or 0),
                        todo=int(item.get("todo") or 0),
                        retest=int(item.get("retest") or 0),
                        blocked=int(item.get("blocked") or 0),
                    )
                exec_key = item.get("key") or ""
                exec_name = item.get("summary") or ""
                dkeys = defects_by_exec.get(exec_key) or []
                exec_ws.append(
                    [
                        exec_key,
                        exec_name,
                        self._milestone_from_execution_name(exec_name),
                        metrics["total"],
                        metrics["coverage"],
                        metrics["coverage_pct"],
                        metrics["passed"],
                        metrics["passed_pct"],
                        metrics["failed"],
                        metrics["failed_pct"],
                        metrics["retest"],
                        metrics["retest_pct"],
                        metrics["untested"],
                        metrics["untested_pct"],
                        metrics["blocked"],
                        metrics["blocked_pct"],
                        len(dkeys),
                        self._cell_list(dkeys),
                        item.get("url") or "",
                    ]
                )
            self._autosize_sheet(exec_ws)

        def_ws = wb.create_sheet("Defects")
        def_ws.append(
            [
                "Defect",
                "Summary",
                "Status",
                "Priority",
                "Linked from failed case",
                "Linked tests",
                "URL",
            ]
        )
        for entry in defect_rows:
            def_ws.append(
                [
                    entry.get("key") or "",
                    entry.get("summary") or "",
                    entry.get("status") or "",
                    entry.get("priority") or "",
                    "Yes" if entry.get("from_failed") else "No",
                    ", ".join(entry.get("linked_tests") or []),
                    entry.get("url") or "",
                ]
            )
        self._autosize_sheet(def_ws)

        # One row per case↔defect link with live Jira status/priority.
        link_ws = wb.create_sheet("Case defects")
        link_ws.append(
            [
                "Test Execution",
                "Test Execution name",
                "Case",
                "Case title",
                "Case status",
                "Defect",
                "Defect summary",
                "Defect status",
                "Defect priority",
                "Defect URL",
            ]
        )
        for entry in case_entries:
            case = entry["case"]
            for key in self._defect_keys(case.get("defects")):
                info = defect_details.get(key) or {}
                link_ws.append(
                    [
                        entry.get("execution_key") or "",
                        entry.get("execution_summary") or "",
                        case.get("key") or "",
                        case.get("summary") or "",
                        case.get("status") or "",
                        key,
                        info.get("summary") or "",
                        info.get("status") or "",
                        info.get("priority") or "",
                        info.get("url") or self.jira.browse_url(key),
                    ]
                )
        self._autosize_sheet(link_ws)

        # Requirements-style rollup by References (RSCDEV), matching coverage report layout.
        ref_ws = wb.create_sheet("References")
        ref_ws.append(
            [
                "refs",
                "Tech Area",
                "Feature",
                "Sub-Feature",
                "Parameters",
                "Total",
                "Coverage",
                "Coverage %",
                "Passed",
                "Passed %",
                "Failed",
                "Failed %",
                "Retest",
                "Retest %",
                "Untested",
                "Untested %",
                "Blocked",
                "Blocked %",
                "Defects",
                "Defect statuses",
                "Defect priorities",
                "URL",
            ]
        )
        cases_by_ref: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for case in all_cases:
            ref_key = self._issue_key_from_reference(case.get("references") or "")
            cases_by_ref[ref_key or "(none)"].append(case)

        for ref_key in sorted(cases_by_ref.keys(), key=lambda k: (k == "(none)", k)):
            ref_cases = cases_by_ref[ref_key]
            metrics = self._metrics_from_cases(ref_cases)
            info = reference_details.get(ref_key) or {}
            defect_keys: list[str] = []
            seen_def: set[str] = set()
            for case in ref_cases:
                for dkey in self._defect_keys(case.get("defects")):
                    if dkey in seen_def:
                        continue
                    seen_def.add(dkey)
                    defect_keys.append(dkey)
            status_bits = []
            priority_bits = []
            for dkey in defect_keys:
                dinfo = defect_details.get(dkey) or {}
                st = (dinfo.get("status") or "").strip()
                pr = (dinfo.get("priority") or "").strip()
                status_bits.append(st)
                priority_bits.append(pr)
            ref_ws.append(
                [
                    ref_key if ref_key != "(none)" else "",
                    info.get("tech_area") or "",
                    info.get("feature") or "",
                    info.get("sub_feature") or "",
                    info.get("parameters") or "",
                    metrics["total"],
                    metrics["coverage"],
                    metrics["coverage_pct"],
                    metrics["passed"],
                    metrics["passed_pct"],
                    metrics["failed"],
                    metrics["failed_pct"],
                    metrics["retest"],
                    metrics["retest_pct"],
                    metrics["untested"],
                    metrics["untested_pct"],
                    metrics["blocked"],
                    metrics["blocked_pct"],
                    self._cell_list(defect_keys),
                    self._cell_list(status_bits, keep_blanks=True),
                    self._cell_list(priority_bits, keep_blanks=True),
                    info.get("url")
                    or (self.jira.browse_url(ref_key) if ref_key != "(none)" else ""),
                ]
            )
        self._autosize_sheet(ref_ws)

        status_ws = wb.create_sheet("Status summary")
        status_ws.append(["Status", "Count", "Percent"])
        for row in summary.chart or []:
            status_ws.append(
                [row.get("status") or "", row.get("count") or 0, row.get("pct") or 0]
            )
        status_ws.append([])
        status_ws.append(["Total", summary.total, ""])
        status_ws.append(["Passed", summary.passed, ""])
        status_ws.append(["Failed", summary.failed, ""])
        status_ws.append(["Todo", summary.todo, ""])
        status_ws.append(["Pass %", summary.pass_pct, ""])
        status_ws.append(["Todo %", summary.todo_pct, ""])
        self._autosize_sheet(status_ws)

        meta = wb.create_sheet("Export info", 0)
        for row in meta_rows:
            meta.append(row)
        meta.append(["Case rows", len(case_entries)])
        meta.append(
            [
                "Failed rows",
                sum(
                    1
                    for e in case_entries
                    if self._normalize_status(e["case"].get("status") or "") == "FAIL"
                ),
            ]
        )
        meta.append(["Defects linked", len(defect_rows)])
        meta.append(["Include manual steps", "Yes" if include_steps else "No"])
        meta.column_dimensions["A"].width = 32
        meta.column_dimensions["B"].width = 56

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def _execution_summary(self, execution_key: str, fallback: str = "") -> str:
        if not execution_key:
            return fallback or ""
        try:
            issue = self.jira.get_issue(execution_key, fields="summary")
            return ((issue.get("fields") or {}).get("summary") or fallback or "").strip()
        except JiraError:
            return fallback or ""

    def export_fail_jira_summary_xlsx(
        self,
        execution_key: str,
        *,
        technology: str | None = None,
        force_refresh: bool = False,
    ) -> bytes:
        """Excel dump: Jira summary for FAIL cases only (same layout as triage).

        Columns: Type, Issue key, No of TCs Impacted, Summary, Status, Priority, Assignee.
        """
        from io import BytesIO

        from openpyxl import Workbook

        if not execution_key:
            raise JiraError("execution is required")

        technology = technology if technology is not None else ""
        cases = self.load_execution_cases(
            execution_key, technology=technology, force_refresh=force_refresh
        )
        fail_cases = [
            c
            for c in cases
            if self._normalize_status(c.get("status") or "") == "FAIL"
        ]

        # Aggregate linked Execution Defects across FAIL cases.
        jira_summary: dict[str, dict[str, Any]] = {}
        all_keys: list[str] = []
        for case in fail_cases:
            case_key = (case.get("key") or "").strip()
            for dkey in self._defect_keys(case.get("defects")):
                canon = dkey.upper()
                entry = jira_summary.setdefault(
                    canon,
                    {
                        "key": dkey,
                        "case_keys": [],
                    },
                )
                if case_key and case_key not in entry["case_keys"]:
                    entry["case_keys"].append(case_key)
                if dkey not in all_keys:
                    all_keys.append(dkey)

        issue_fields = (
            self._fetch_defect_issue_fields(all_keys) if all_keys else {}
        )

        wb = Workbook()
        meta = wb.active
        meta.title = "Export info"
        meta.append(["Scope", "FAIL cases — Jira summary (triage-style)"])
        meta.append(["Test Execution", execution_key])
        meta.append(
            ["Execution name", self._execution_summary(execution_key)]
        )
        meta.append(["Technology filter", technology or "(all)"])
        meta.append(["FAIL cases in run", len(fail_cases)])
        meta.append(["Unique linked Jiras", len(jira_summary)])
        self._autosize_sheet(meta)

        summary_ws = wb.create_sheet("Jira summary")
        summary_ws.append(
            [
                "Type",
                "Issue key",
                "No of TCs Impacted",
                "Summary",
                "Status",
                "Priority",
                "Assignee",
            ]
        )
        total_impacted = 0
        for entry in sorted(
            jira_summary.values(),
            key=lambda e: (-len(e.get("case_keys") or []), e.get("key") or ""),
        ):
            jira_key = entry.get("key") or ""
            info = issue_fields.get(jira_key) or issue_fields.get(jira_key.upper()) or {}
            if not info:
                for ik, iv in issue_fields.items():
                    if (ik or "").upper() == jira_key.upper():
                        info = iv
                        break
            impacted = len(entry.get("case_keys") or [])
            total_impacted += impacted
            summary_ws.append(
                [
                    (info.get("issuetype") or "Bug").strip() or "Bug",
                    jira_key,
                    impacted,
                    (info.get("summary") or "").strip(),
                    (info.get("status") or "").strip(),
                    (info.get("priority") or "").strip(),
                    (info.get("assignee") or "").strip(),
                ]
            )
        summary_ws.append(["Total", "", total_impacted, "", "", "", ""])
        self._autosize_sheet(summary_ws)

        # Optional case→jira mapping sheet (compact).
        map_ws = wb.create_sheet("FAIL cases")
        map_ws.append(
            [
                "Case",
                "Title",
                "Section",
                "Linked Jiras",
                "URL",
            ]
        )
        for case in sorted(fail_cases, key=lambda c: c.get("key") or ""):
            keys = self._defect_keys(case.get("defects"))
            map_ws.append(
                [
                    case.get("key") or "",
                    case.get("summary") or "",
                    case.get("section_path") or "",
                    self._cell_list(keys),
                    case.get("url") or "",
                ]
            )
        self._autosize_sheet(map_ws)

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def export_execution_xlsx(
        self,
        execution_key: str,
        *,
        section_path: str = "",
        status_filter: str = "",
        search: str = "",
        technology: str | None = None,
        force_refresh: bool = False,
        include_steps: bool = True,
        execution_summary: str = "",
    ) -> bytes:
        """Build an .xlsx workbook of cases for a Test Execution (optional filters)."""
        if not execution_key:
            raise JiraError("execution is required")

        technology = technology if technology is not None else ""
        exec_summary = execution_summary or self._execution_summary(execution_key)
        cases = self.load_execution_cases(
            execution_key, technology=technology, force_refresh=force_refresh
        )
        filtered = self._filter_cases(cases, section_path, status_filter, search)
        case_entries = [
            {
                "case": case,
                "execution_key": execution_key,
                "execution_summary": exec_summary,
            }
            for case in filtered
        ]
        return self._write_export_workbook(
            meta_rows=[
                ["Scope", "Single Test Execution"],
                ["Execution", execution_key],
                ["Execution name", exec_summary],
                ["Technology filter", technology or "(all)"],
                ["Section", section_path or "(all)"],
                ["Status filter", status_filter or "(all)"],
                ["Search", search or ""],
                ["Total in run (after tech filter)", len(cases)],
            ],
            case_entries=case_entries,
            include_steps=include_steps,
        )

    def export_plan_xlsx(
        self,
        plan_key: str,
        *,
        technology: str | None = None,
        force_refresh: bool = False,
        include_steps: bool = False,
        progress_cb: Any | None = None,
        is_cancelled: Any | None = None,
    ) -> bytes:
        """Dump all Test Executions under a Test Plan into one workbook."""
        from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, wait

        from .export_jobs import ExportCancelled

        def _cancelled() -> bool:
            return bool(callable(is_cancelled) and is_cancelled())

        def report(**payload: Any) -> None:
            if _cancelled():
                raise ExportCancelled("Export cancelled by user")
            if callable(progress_cb):
                progress_cb(payload)

        if not plan_key:
            raise JiraError("plan is required")

        technology = technology if technology is not None else ""
        report(
            status="running",
            phase="listing",
            message="Listing Test Executions…",
            current="",
            done=0,
            total=0,
        )
        if _cancelled():
            raise ExportCancelled("Export cancelled by user")
        raw_execs = self.xray.get_test_plan_executions(plan_key)
        if _cancelled():
            raise ExportCancelled("Export cancelled by user")
        executions: list[dict[str, str]] = []
        for item in raw_execs or []:
            if not isinstance(item, dict):
                continue
            key = (item.get("key") or "").strip()
            if not key:
                continue
            executions.append(
                {
                    "key": key,
                    "summary": (item.get("summary") or "").strip(),
                }
            )
        if not executions:
            raise JiraError(f"No Test Executions linked to {plan_key}")
        if _cancelled():
            raise ExportCancelled("Export cancelled by user")

        report(
            phase="loading",
            message=(
                f"Loading 0/{len(executions)} Test Executions "
                f"(up to 4 in parallel)…"
            ),
            total=len(executions),
            done=0,
            current="",
        )

        def _load_one(exec_info: dict[str, str]) -> dict[str, Any]:
            if _cancelled():
                raise ExportCancelled("Export cancelled by user")
            key = exec_info["key"]
            summary = exec_info.get("summary") or self._execution_summary(key)
            if _cancelled():
                raise ExportCancelled("Export cancelled by user")
            cases = self.load_execution_cases(
                key, technology=technology, force_refresh=force_refresh
            )
            if _cancelled():
                raise ExportCancelled("Export cancelled by user")
            metrics = self._metrics_from_cases(cases)
            return {
                "key": key,
                "summary": summary,
                "cases": cases,
                "metrics": metrics,
                "total": metrics["total"],
                "passed": metrics["passed"],
                "failed": metrics["failed"],
                "todo": metrics["untested"],
                "coverage": metrics["coverage"],
                "pass_pct": metrics["passed_pct"],
                "fail_pct": metrics["failed_pct"],
                "todo_pct": metrics["untested_pct"],
                "coverage_pct": metrics["coverage_pct"],
                "retest": metrics["retest"],
                "blocked": metrics["blocked"],
                "url": self.jira.browse_url(key),
            }

        loaded: list[dict[str, Any]] = []
        errors: list[str] = []
        workers = min(4, max(1, len(executions)))
        done_count = 0
        pool = ThreadPoolExecutor(max_workers=workers)
        futures = {pool.submit(_load_one, ex): ex["key"] for ex in executions}
        pending = set(futures.keys())
        try:
            while pending:
                if _cancelled():
                    for fut in pending:
                        fut.cancel()
                    # Do not wait for in-flight HTTP loads — abandon immediately.
                    raise ExportCancelled("Export cancelled by user")
                done_set, pending = wait(
                    pending, timeout=0.2, return_when=FIRST_COMPLETED
                )
                if not done_set:
                    continue
                for fut in done_set:
                    key = futures.get(fut) or ""
                    try:
                        loaded.append(fut.result())
                    except ExportCancelled:
                        for left in pending:
                            left.cancel()
                        raise
                    except Exception as exc:  # noqa: BLE001
                        errors.append(f"{key}: {exc}")
                    done_count += 1
                    report(
                        phase="loading",
                        message=(
                            f"Loading {done_count}/{len(executions)} "
                            f"Test Executions…"
                        ),
                        total=len(executions),
                        done=done_count,
                        current=key,
                    )
        except ExportCancelled:
            for fut in pending:
                fut.cancel()
            raise
        finally:
            # wait=False so Cancel returns immediately even if a run is mid-fetch.
            pool.shutdown(wait=False, cancel_futures=True)

        if _cancelled():
            raise ExportCancelled("Export cancelled by user")

        loaded.sort(key=lambda x: x.get("key") or "")
        case_entries: list[dict[str, Any]] = []
        for item in loaded:
            for case in item.get("cases") or []:
                case_entries.append(
                    {
                        "case": case,
                        "execution_key": item.get("key") or "",
                        "execution_summary": item.get("summary") or "",
                    }
                )

        report(
            phase="building",
            message="Building Excel workbook…",
            total=len(executions),
            done=len(executions),
            current="",
        )
        plan_summary = self._execution_summary(plan_key)
        meta = [
            ["Scope", "All Test Executions in Test Plan"],
            ["Plan", plan_key],
            ["Plan name", plan_summary],
            ["Technology filter", technology or "(all)"],
            ["Executions", len(loaded)],
            ["Execution load errors", "; ".join(errors) if errors else ""],
        ]
        return self._write_export_workbook(
            meta_rows=meta,
            case_entries=case_entries,
            include_steps=include_steps,
            executions_sheet=loaded,
        )

    @staticmethod
    def _pct(part: int, total: int) -> float:
        return round((part / total) * 100, 1) if total else 0.0

    def _metrics_from_counts(
        self,
        *,
        total: int,
        passed: int = 0,
        failed: int = 0,
        todo: int = 0,
        retest: int = 0,
        blocked: int = 0,
    ) -> dict[str, Any]:
        coverage = passed + failed
        return {
            "total": total,
            "coverage": coverage,
            "coverage_pct": self._pct(coverage, total),
            "passed": passed,
            "passed_pct": self._pct(passed, total),
            "failed": failed,
            "failed_pct": self._pct(failed, total),
            "retest": retest,
            "retest_pct": self._pct(retest, total),
            "untested": todo,
            "untested_pct": self._pct(todo, total),
            "blocked": blocked,
            "blocked_pct": self._pct(blocked, total),
        }

    def _metrics_from_cases(self, cases: list[dict[str, Any]]) -> dict[str, Any]:
        counter: Counter[str] = Counter()
        for case in cases:
            counter[self._normalize_status(case.get("status") or "TODO")] += 1
        total = sum(counter.values())
        return self._metrics_from_counts(
            total=total,
            passed=counter.get("PASS", 0),
            failed=counter.get("FAIL", 0),
            todo=counter.get("TODO", 0),
            retest=counter.get("RETEST", 0),
            blocked=counter.get("BLOCKED", 0),
        )

    def _summarize_statuses(self, cases: list[dict[str, Any]]) -> StatusSummary:
        counter: Counter[str] = Counter()
        for case in cases:
            counter[self._normalize_status(case.get("status") or "TODO")] += 1

        total = sum(counter.values())
        passed = counter.get("PASS", 0)
        failed = counter.get("FAIL", 0)
        todo = counter.get("TODO", 0)
        pass_pct = round((passed / total) * 100, 1) if total else 0.0
        todo_pct = round((todo / total) * 100, 1) if total else 0.0

        order = ["PASS", "FAIL", "BLOCKED", "RETEST", "TODO", "EXECUTING", "ABORTED", "NA"]
        chart = []
        for status in order:
            if counter.get(status):
                chart.append(
                    {
                        "status": status,
                        "count": counter[status],
                        "pct": round((counter[status] / total) * 100, 1) if total else 0,
                        "color": STATUS_COLORS.get(status, STATUS_COLORS["OTHER"]),
                    }
                )
        for status, count in counter.items():
            if status not in order:
                chart.append(
                    {
                        "status": status,
                        "count": count,
                        "pct": round((count / total) * 100, 1) if total else 0,
                        "color": STATUS_COLORS.get(status, STATUS_COLORS["OTHER"]),
                    }
                )

        return StatusSummary(
            counts=dict(counter),
            total=total,
            passed=passed,
            failed=failed,
            todo=todo,
            pass_pct=pass_pct,
            todo_pct=todo_pct,
            chart=chart,
        )

    def _fetch_defect_issue_fields(
        self, keys: list[str]
    ) -> dict[str, dict[str, Any]]:
        """Fetch summary/status/priority/description/people for defect issue keys."""
        out: dict[str, dict[str, Any]] = {}
        unique = [k for k in dict.fromkeys(keys) if k and self._looks_like_issue_key(k)]
        if not unique:
            return out
        batch_size = 50
        fields = [
            "summary",
            "status",
            "priority",
            "issuetype",
            "description",
            "assignee",
            "reporter",
            "created",
            "updated",
        ]

        def _pack(issue: dict[str, Any], fallback_key: str = "") -> dict[str, Any]:
            f = issue.get("fields") or {}
            issue_type = f.get("issuetype") or {}
            type_name = ""
            if isinstance(issue_type, dict):
                type_name = (issue_type.get("name") or "").strip()
            elif issue_type:
                type_name = str(issue_type).strip()
            return {
                "summary": f.get("summary") or "",
                "status": ((f.get("status") or {}).get("name") or ""),
                "priority": ((f.get("priority") or {}).get("name") or ""),
                "issuetype": type_name or "Bug",
                "description": self._plain_description(f.get("description")),
                "assignee": self._user_name(f.get("assignee")),
                "reporter": self._user_name(f.get("reporter")),
                "created": f.get("created"),
                "updated": f.get("updated"),
            }

        for i in range(0, len(unique), batch_size):
            batch = unique[i : i + batch_size]
            quoted = ", ".join(f'"{k}"' for k in batch)
            try:
                issues = self.jira.search_all(
                    jql=f"key in ({quoted})",
                    fields=fields,
                    page_size=min(100, len(batch)),
                    hard_limit=len(batch) + 5,
                )
            except JiraError:
                # Fall back to single-issue reads for this batch.
                for key in batch:
                    try:
                        issue = self.jira.get_issue(key, fields=",".join(fields))
                    except JiraError:
                        continue
                    out[issue.get("key") or key] = _pack(issue, key)
                continue
            for issue in issues or []:
                key = issue.get("key") or ""
                if not key:
                    continue
                out[key] = _pack(issue, key)
        return out

    def _collect_defects(self, cases: list[dict[str, Any]]) -> list[dict[str, Any]]:
        defect_map: dict[str, dict[str, Any]] = {}
        for case in cases:
            for defect in case.get("defects") or []:
                key = defect if isinstance(defect, str) else (defect or {}).get("key")
                if not key:
                    continue
                key = str(key).strip()
                entry = defect_map.setdefault(
                    key,
                    {
                        "key": key,
                        "url": self.jira.browse_url(key),
                        "linked_tests": [],
                        "from_failed": False,
                        "summary": "",
                        "status": "",
                        "priority": "",
                    },
                )
                case_key = case.get("key") or ""
                if case_key and case_key not in entry["linked_tests"]:
                    entry["linked_tests"].append(case_key)
                if self._normalize_status(case.get("status") or "") == "FAIL":
                    entry["from_failed"] = True

        enriched = self._fetch_defect_issue_fields(list(defect_map.keys()))
        for key, fields in enriched.items():
            if key in defect_map:
                defect_map[key].update(fields)

        return sorted(
            defect_map.values(),
            key=lambda d: (-len(d.get("linked_tests") or []), d["key"]),
        )

    def _build_coverage(self, cases: list[dict[str, Any]]) -> dict[str, Any]:
        # Approximate coverage using section folders as requirement/feature buckets
        by_section: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for case in cases:
            path = self._clean_path(case.get("section_path") or "Uncategorized")
            top = path.split("/")[0] if path else "Uncategorized"
            by_section[top].append(case)

        rows = []
        for section, section_cases in sorted(by_section.items()):
            summary = self._summarize_statuses(section_cases)
            covered = summary.total - summary.todo
            rows.append(
                {
                    "section": section,
                    "total": summary.total,
                    "executed": covered,
                    "passed": summary.passed,
                    "failed": summary.failed,
                    "todo": summary.todo,
                    "coverage_pct": round((covered / summary.total) * 100, 1)
                    if summary.total
                    else 0,
                    "pass_pct": summary.pass_pct,
                }
            )
        return {
            "rows": rows,
            "total_sections": len(rows),
            "total_tests": len(cases),
        }

    @staticmethod
    def _looks_like_issue_key(value: str) -> bool:
        import re

        return bool(re.fullmatch(r"[A-Za-z][A-Za-z0-9_]+-\d+", (value or "").strip()))

    @staticmethod
    def _normalize_status(value: str) -> str:
        key = (value or "TODO").strip().upper()
        return STATUS_ALIASES.get(key, key if key else "TODO")

    @staticmethod
    def _clean_path(path: str) -> str:
        path = (path or "").replace("\\", "/").strip().strip("/")
        parts = [p.strip() for p in path.split("/") if p.strip()]
        return "/".join(parts)

    def _enrich_run_assignees(self, rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
        """Fill missing assignees via GET /testrun/{id}/assignee (Xray column source)."""
        from concurrent.futures import ThreadPoolExecutor, as_completed

        need: list[tuple[int, dict[str, Any]]] = []
        for idx, item in enumerate(rows):
            if not isinstance(item, dict):
                continue
            if self._run_assignee(item):
                continue
            run_id = item.get("id") or item.get("testRunId")
            if run_id:
                need.append((idx, item))
        if not need:
            return rows

        def _one(pair: tuple[int, dict[str, Any]]) -> tuple[int, Any]:
            idx, item = pair
            run_id = item.get("id") or item.get("testRunId")
            try:
                return idx, self.xray.get_test_run_assignee(run_id)
            except JiraError:
                return idx, None

        with ThreadPoolExecutor(max_workers=10) as pool:
            futures = [pool.submit(_one, pair) for pair in need]
            for fut in as_completed(futures):
                idx, user = fut.result()
                if user is None:
                    continue
                rows[idx] = {**rows[idx], "assignee": user}
        return rows

    @classmethod
    def _run_assignee_info(cls, item: dict[str, Any] | None) -> tuple[str, str]:
        """Return (displayName, userKey) for a Test Run assignee payload."""
        if not isinstance(item, dict):
            return "", ""
        for key in (
            "assignee",
            "assignedTo",
            "assigned_to",
            "user",
            "testedBy",
            "executor",
        ):
            raw = item.get(key)
            display = cls._user_name(raw)
            user_key = cls._user_key(raw)
            if display or user_key:
                return display or user_key, user_key or display
        nested = item.get("test") if isinstance(item.get("test"), dict) else None
        if nested:
            for key in ("assignee", "assignedTo", "user"):
                raw = nested.get(key)
                display = cls._user_name(raw)
                user_key = cls._user_key(raw)
                if display or user_key:
                    return display or user_key, user_key or display
        return "", ""

    @classmethod
    def _run_assignee(cls, item: dict[str, Any] | None) -> str:
        """Extract display name for a Test Run assignee from Xray payloads."""
        return cls._run_assignee_info(item)[0]

    @staticmethod
    def _user_key(user: Any) -> str:
        if not user:
            return ""
        if isinstance(user, str):
            return user.strip()
        if not isinstance(user, dict):
            return str(user).strip()
        return (
            user.get("name")
            or user.get("key")
            or user.get("accountId")
            or user.get("username")
            or ""
        ).strip()

    @staticmethod
    def _user_name(user: Any) -> str:
        if not user:
            return ""
        if isinstance(user, str):
            return user.strip()
        if not isinstance(user, dict):
            return str(user).strip()
        return (
            user.get("displayName")
            or user.get("display_name")
            or user.get("name")
            or user.get("key")
            or user.get("accountId")
            or user.get("emailAddress")
            or ""
        ).strip()

    @classmethod
    def _plain_description(cls, value: Any) -> str:
        """Normalize Jira description (wiki string or ADF-like dict) to plain text."""
        if value is None:
            return ""
        if isinstance(value, str):
            return value.strip()
        if isinstance(value, dict):
            # Atlassian Document Format (Cloud) or wrapped content
            if "content" in value:
                return cls._adf_to_text(value).strip()
            return (
                value.get("content")
                or value.get("value")
                or value.get("text")
                or ""
            )
        return str(value).strip()

    @classmethod
    def _adf_to_text(cls, node: Any) -> str:
        if node is None:
            return ""
        if isinstance(node, str):
            return node
        if isinstance(node, list):
            return "".join(cls._adf_to_text(n) for n in node)
        if not isinstance(node, dict):
            return str(node)
        node_type = node.get("type") or ""
        text = node.get("text") or ""
        content = node.get("content") or []
        inner = cls._adf_to_text(content) if content else text
        if node_type in {"paragraph", "heading", "blockquote", "listItem", "codeBlock"}:
            return f"{inner}\n"
        if node_type == "hardBreak":
            return "\n"
        return inner

    @staticmethod
    def _extract_field(fields: dict[str, Any], field_id: str | None) -> Any:
        if not field_id:
            return ""
        value = fields.get(field_id)
        if value is None:
            return ""
        if isinstance(value, str):
            return value
        if isinstance(value, dict):
            return (
                value.get("name")
                or value.get("value")
                or value.get("key")
                or value.get("displayName")
                or ""
            )
        if isinstance(value, list):
            parts = []
            for item in value:
                if isinstance(item, str):
                    parts.append(item)
                elif isinstance(item, dict):
                    parts.append(
                        item.get("name")
                        or item.get("value")
                        or item.get("key")
                        or item.get("displayName")
                        or ""
                    )
            return ", ".join([p for p in parts if p])
        return str(value)

    @staticmethod
    def _path_from_labels(labels: list[str]) -> str:
        if not labels:
            return ""
        # Prefer hierarchical-looking labels; else join distinctive ones
        interesting = [l for l in labels if l and l.lower() not in {"functional", "deployed"}]
        if len(interesting) >= 2:
            return "/".join(interesting[:3])
        return interesting[0] if interesting else labels[0]

    @staticmethod
    def _defect_keys_from_links(links: list[dict[str, Any]]) -> list[str]:
        keys = []
        for link in links:
            for side in ("outwardIssue", "inwardIssue"):
                issue = link.get(side)
                if not issue:
                    continue
                issue_type = ((issue.get("fields") or {}).get("issuetype") or {}).get("name", "")
                key = issue.get("key")
                if key and ("bug" in issue_type.lower() or "defect" in issue_type.lower()):
                    keys.append(key)
                elif key and side == "outwardIssue":
                    # keep linked issues as potential defects when type missing in search payload
                    keys.append(key)
        return list(dict.fromkeys(keys))


def get_service() -> DashboardService:
    # Fresh client each call so session/env credentials are picked up.
    return DashboardService()
